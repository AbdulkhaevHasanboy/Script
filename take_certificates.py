#!/usr/bin/env python3
import os
import sys
import csv
import json
import subprocess
import openpyxl
from pathlib import Path

# Paths
EXCEL_PATH = Path("Names.xlsx")
CSV_PATH = Path("students.csv")
RESULTS_PATH = Path("results.json")

def normalize_email(email_str):
    email_str = email_str.strip().lower()
    if "@gmail.com" in email_str:
        username, domain = email_str.split("@")
        return username.replace(".", "") + "@" + domain
    return email_str

def get_student_index_in_csv(passport):
    students = []
    headers = ["student_id", "first_name", "last_name", "email", "certificate_url", "password"]
    
    if CSV_PATH.exists():
        try:
            with open(CSV_PATH, "r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or headers
                students = list(reader)
        except Exception as e:
            print(f"Error reading students.csv: {e}")

    for idx, s in enumerate(students):
        if s.get("password") == passport or s.get("student_id") == passport:
            return idx + 1 # 1-based index
    return None

def update_or_add_student_in_csv(full_name, passport, email_addr):
    students = []
    headers = ["student_id", "first_name", "last_name", "email", "certificate_url", "password"]
    
    if CSV_PATH.exists():
        try:
            with open(CSV_PATH, "r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or headers
                students = list(reader)
        except Exception as e:
            print(f"Error reading students.csv: {e}")
            
    # Check if student exists
    found_idx = -1
    for idx, s in enumerate(students):
        if s.get("password") == passport:
            found_idx = idx
            break
            
    parts = full_name.split()
    first_name = parts[0] if parts else ""
    last_name = " ".join(parts[1:]) if len(parts) > 1 else ""

    student_row = {
        "student_id": passport,
        "first_name": first_name,
        "last_name": last_name,
        "email": email_addr,
        "certificate_url": students[found_idx].get("certificate_url", "") if found_idx != -1 else "",
        "password": passport
    }
    
    for h in headers:
        if h not in student_row:
            student_row[h] = ""
            
    if found_idx != -1:
        students[found_idx] = student_row
        student_index = found_idx + 1
    else:
        students.append(student_row)
        student_index = len(students)
        
    try:
        with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(students)
    except Exception as e:
        print(f"Error writing students.csv: {e}")
        
    return student_index

def check_if_cert_obtained(passport):
    if not CSV_PATH.exists():
        return None
    try:
        with open(CSV_PATH, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for s in reader:
                if s.get("password") == passport:
                    cert = s.get("certificate_url", "").strip()
                    return cert if cert else None
    except Exception:
        pass
    return None

def run_replay_for_student(index, full_name, invite_url):
    print(f"\n🚀 Starting automated Playwright browser session for student: {full_name}")
    print(f"🔗 Coursera Link: {invite_url}")
    print(f"Index in CSV: {index}\n" + "-"*50)
    
    env = os.environ.copy()
    env["START"] = str(index)
    env["END"] = str(index)
    env["CONCURRENCY"] = "1"
    env["COURSE_URL"] = invite_url
    env["MODE"] = "auto"
    
    cmd = ["node", "coursera_manual_runner.js"]
    try:
        process = subprocess.Popen(cmd, env=env, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
        for line in process.stdout:
            sys.stdout.write(line)
            sys.stdout.flush()
        process.wait()
        print("-"*50 + f"\nPlaywright session exited with code {process.returncode}.")
        return process.returncode == 0
    except Exception as e:
        print(f"❌ Playwright session failed to execute: {e}")
        return False

def main():
    print("=" * 60)
    print("Coursera Certificate Fetcher Script Started")
    print("=" * 60)

    if not EXCEL_PATH.exists():
        print(f"Error: {EXCEL_PATH} not found.")
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet = wb["Talabalar"]
    except Exception as e:
        print(f"Error loading Names.xlsx: {e}")
        return

    # Load results.json to check/save
    results = []
    if RESULTS_PATH.exists():
        try:
            results = json.loads(RESULTS_PATH.read_text())
        except Exception:
            pass

    processed_count = 0

    for r in range(3, sheet.max_row + 1):
        full_name = str(sheet.cell(r, 1).value or "").strip()
        passport = str(sheet.cell(r, 2).value or "").strip()
        invite_url = str(sheet.cell(r, 6).value or "").strip()
        email_addr = str(sheet.cell(r, 7).value or "").strip()
        status_val = str(sheet.cell(r, 8).value or "").strip()

        if not passport:
            continue

        # Check if certificate is already resolved in Column H or results.json
        has_cert = False
        existing_cert = None
        
        if "coursera.org/share" in status_val or "verify" in status_val:
            has_cert = True
            existing_cert = status_val
        else:
            for entry in results:
                if entry.get("document") == passport and entry.get("certificate_url"):
                    has_cert = True
                    existing_cert = entry.get("certificate_url")
                    break

        # Pass condition: No invite link yet OR certificate already exists
        if not invite_url or "coursera.org" not in invite_url:
            # Just pass (skip)
            continue

        if has_cert:
            # Just pass (skip)
            continue

        print(f"\n[Row {r}] Processing: {full_name}")
        
        # Add to CSV and get index
        student_index = update_or_add_student_in_csv(full_name, passport, email_addr)

        # Run replayer
        run_replay_for_student(student_index, full_name, invite_url)

        # Check success
        cert_url = check_if_cert_obtained(passport)
        if cert_url:
            print(f"✅ Success! Captured certificate: {cert_url}")
            
            # Write to Excel Column H
            sheet.cell(r, 8, value=cert_url)
            wb.save(EXCEL_PATH)
            
            # Write to results.json
            found_in_results = False
            for entry in results:
                if entry.get("document") == passport:
                    entry["certificate_url"] = cert_url
                    entry["activated"] = True
                    found_in_results = True
                    break
            if not found_in_results:
                results.append({
                    "row": r,
                    "document": passport,
                    "email": email_addr,
                    "certificate_url": cert_url,
                    "activated": True
                })
            RESULTS_PATH.write_text(json.dumps(results, ensure_ascii=False, indent=2))
            processed_count += 1
        else:
            print(f"❌ Failed to obtain certificate for {full_name}.")

    print(f"\nFinished. Processed {processed_count} new certificates.")

if __name__ == "__main__":
    main()
