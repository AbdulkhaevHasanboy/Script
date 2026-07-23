import imaplib
import quopri
import re
import json
import email
import time
import openpyxl
import requests
from pathlib import Path
from email.header import decode_header
import concurrent.futures

GMAIL_USER = "qwertyuioplkjhgfdsazxcvbnmhrh@gmail.com"
GMAIL_APP_PASS = "kzpsmykjzrueegpv"

EXCEL_PATH = Path("Names.xlsx")
DB_PATH = Path("Names_db.json")

def normalize_name(name):
    if not name:
        return ""
    name = str(name).strip().upper()
    name = re.sub(r"['`‘’ʻʼ]", "", name)
    name = re.sub(r"\s+", " ", name)
    return name

def normalize_email_for_match(addr):
    if not addr or "@" not in addr:
        return ""
    parts = addr.lower().split("@")
    user = parts[0].replace(".", "")
    return f"{user}@{parts[1]}"

def get_decoded_header(header_val):
    if not header_val:
        return ""
    decoded_parts = decode_header(header_val)
    header_text = ""
    for part, encoding in decoded_parts:
        if isinstance(part, bytes):
            try:
                header_text += part.decode(encoding or "utf-8", errors="ignore")
            except Exception:
                header_text += part.decode("utf-8", errors="ignore")
        else:
            header_text += str(part)
    return header_text

def resolve_redirect(url):
    try:
        r = requests.head(url, allow_redirects=True, timeout=8)
        if r.url and ("coursera.org" in r.url or "invitationToken=" in r.url):
            return r.url
    except Exception:
        pass
    try:
        r = requests.get(url, allow_redirects=True, timeout=8, stream=True)
        if r.url and ("coursera.org" in r.url or "invitationToken=" in r.url):
            return r.url
    except Exception:
        pass
    return url

def main():
    print("📖 Loading student records from Names.xlsx and Names_db.json...")
    db = []
    if DB_PATH.exists():
        try:
            db = json.loads(DB_PATH.read_text())
        except Exception:
            pass

    # Build student lookups
    email_map = {}
    students_parsed = []

    for r in db:
        passport = str(r.get("document") or "").strip()
        name = str(r.get("name") or "").strip()
        email_addr = str(r.get("email") or "").strip().lower()
        invite_url = str(r.get("invite_url") or "").strip()
        if invite_url and "coursera.org" in invite_url:
            continue

        norm_email = normalize_email_for_match(email_addr)
        norm_name = normalize_name(name)
        name_words = tuple(norm_name.split())

        s_obj = {
            "row": r.get("row"),
            "passport": passport,
            "name": name,
            "email": email_addr,
            "norm_email": norm_email,
            "norm_name_words": set(name_words),
            "invite_url": invite_url
        }

        if norm_email:
            email_map[norm_email] = s_obj
        if name_words:
            students_parsed.append(s_obj)

    print(f"Loaded {len(students_parsed)} students from database.")

    # Connect to Gmail IMAP with retry loop to handle startup quota limits
    mail = None
    reconnect_delay = 30
    while True:
        try:
            print("Connecting to Gmail IMAP...", flush=True)
            mail = imaplib.IMAP4_SSL("imap.gmail.com", timeout=15)
            mail.login(GMAIL_USER, GMAIL_APP_PASS)
            
            print("📂 Scanning INBOX for Coursera invitation emails...", flush=True)
            status, _ = mail.select('"INBOX"', readonly=True)
            if status != "OK":
                raise Exception("Failed to select INBOX")
                
            status, data = mail.search(None, '(SINCE "15-Jul-2026" SUBJECT "invited")')
            if status == "OK" and data[0]:
                break
            else:
                print("No invitation emails found or search failed. Retrying search...")
                raise Exception("Search failed or empty")
        except Exception as e:
            print(f"⚠️ Initial connection/search failed ({e}). Retrying in {reconnect_delay}s...", flush=True)
            if mail:
                try:
                    mail.logout()
                except Exception:
                    pass
            time.sleep(reconnect_delay)
            reconnect_delay = min(reconnect_delay + 10, 60)

    msg_ids = data[0].split()
    print(f"Found {len(msg_ids)} total invitation emails in INBOX.", flush=True)

    # Fetch bodies in batches of 35
    batch_size = 35
    total_batches = (len(msg_ids) + batch_size - 1) // batch_size
    extracted_unresolved = {}  # row -> link

    for i in range(0, len(msg_ids), batch_size):
        batch = msg_ids[i:i+batch_size]
        batch_str = b",".join(batch)
        batch_idx = i // batch_size + 1

        print(f"  ✉️ Fetching bodies batch {batch_idx}/{total_batches}...", flush=True)
        
        body_data = None
        for attempt in range(1, 4):
            try:
                status, body_data = mail.fetch(batch_str, "(BODY.PEEK[TEXT])")
                if status == "OK" and body_data:
                    break
            except Exception as e:
                err_str = str(e)
                print(f"  ⚠️ Fetch error: {err_str}", flush=True)
                
                # Close current connection if it's broken or errored
                try:
                    mail.logout()
                except Exception:
                    pass
                
                print("  🔄 Reconnecting to Gmail IMAP...", flush=True)
                # Keep retrying reconnection until successful to prevent illegal state errors
                reconnect_delay = 30
                while True:
                    try:
                        time.sleep(reconnect_delay)
                        mail = imaplib.IMAP4_SSL("imap.gmail.com", timeout=15)
                        mail.login(GMAIL_USER, GMAIL_APP_PASS)
                        status, _ = mail.select('"INBOX"', readonly=True)
                        if status == "OK":
                            print("  ✅ Reconnection successful.", flush=True)
                            break
                    except Exception as conn_err:
                        print(f"  ⚠️ Reconnection failed ({conn_err}). Retrying in {reconnect_delay}s...", flush=True)
                        reconnect_delay = min(reconnect_delay + 10, 60)

        if not body_data:
            continue

        for part in body_data:
            if isinstance(part, tuple):
                try:
                    decoded_bytes = quopri.decodestring(part[1])
                    body_text = decoded_bytes.decode("utf-8", errors="ignore")

                    # 1. Extract link
                    button_pattern = r'(?:Join now|Join Program|Start learning|Accept invitation|Accept Invitation)\s*<(https://[^\s"<>]+coursera\.org/[^>]+)>'
                    match = re.search(button_pattern, body_text, re.IGNORECASE)
                    if not match:
                        continue
                    
                    l = match.group(1)
                    parts_url = l.split("?")
                    invite_link = parts_url[0] + "?" + parts_url[1].split()[0] if len(parts_url) > 1 else l
                    invite_link = invite_link.rstrip('>').rstrip('"').rstrip(')').rstrip('.')

                    # 2. Extract student name from greeting
                    target_student = None
                    m_name = re.search(r'Hello\s+([^,\n\r]+),', body_text, re.IGNORECASE)
                    if m_name:
                        extracted_name = m_name.group(1).strip()
                        extracted_words = set(normalize_name(extracted_name).split())

                        if len(extracted_words) >= 2:
                            for s in students_parsed:
                                if extracted_words.issubset(s["norm_name_words"]):
                                    target_student = s
                                    break

                    if target_student:
                        extracted_unresolved[target_student["row"]] = {
                            "link": invite_link,
                            "student": target_student
                        }
                except Exception:
                    pass

    mail.logout()
    print(f"\nSuccessfully extracted {len(extracted_unresolved)} invitation links! Resolving redirects in parallel...", flush=True)

    # Resolve redirects with 20 parallel workers
    updates_to_save = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
        future_to_row = {
            executor.submit(resolve_redirect, item["link"]): (row, item["student"])
            for row, item in extracted_unresolved.items()
        }
        for future in concurrent.futures.as_completed(future_to_row):
            row, student = future_to_row[future]
            try:
                final_url = future.result()
                if "coursera.org" in final_url:
                    updates_to_save[row] = {
                        "invite_link": final_url,
                        "email": student["email"],
                        "passport": student["passport"],
                        "name": student["name"]
                    }
                    print(f"  ✨ Resolved Row {row}: {student['name']} -> {final_url[:70]}...", flush=True)
            except Exception as e:
                print(f"⚠️ Error resolving row {row}: {e}", flush=True)

    print(f"\n💾 Saving {len(updates_to_save)} resolved links to Names.xlsx and Names_db.json...", flush=True)

    # Update Names_db.json
    for r in db:
        row_num = r.get("row")
        if row_num in updates_to_save:
            r["invite_url"] = updates_to_save[row_num]["invite_link"]
            r["activated"] = True

    DB_PATH.write_text(json.dumps(db, ensure_ascii=False, indent=2))
    print("  ✅ Names_db.json updated.")

    # Update Names.xlsx
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Talabalar"]
        ws.cell(row=1, column=6, value="Invite Link")
        ws.cell(row=1, column=7, value="Email")

        for r_idx in range(2, ws.max_row + 1):
            pass_val = str(ws.cell(row=r_idx, column=2).value or "").strip()
            for row_num, info in updates_to_save.items():
                if info["passport"] and pass_val == info["passport"]:
                    ws.cell(row=r_idx, column=6, value=info["invite_link"])
                    if info["email"]:
                        ws.cell(row=r_idx, column=7, value=info["email"])
                    break

        wb.save(EXCEL_PATH)
        print("  ✅ Names.xlsx updated.")

    print(f"\nDone! Successfully synchronized {len(updates_to_save)} invitation links.")

if __name__ == "__main__":
    main()
