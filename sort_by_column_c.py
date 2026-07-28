#!/usr/bin/env python3
import csv
import fcntl
import openpyxl
from pathlib import Path

NEW_CSV_PATH = Path("NEW.csv")
EXCEL_PATH = Path("Names.xlsx")

def sort_new_csv_by_column_c():
    if not NEW_CSV_PATH.exists():
        print(f"❌ {NEW_CSV_PATH} not found.")
        return

    lock_path = NEW_CSV_PATH.with_suffix(".csv.lock")
    with open(lock_path, "w") as lock_f:
        fcntl.flock(lock_f, fcntl.LOCK_EX)
        
        with open(NEW_CSV_PATH, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        header = rows[0]
        data_rows = rows[1:]

        # Sort data rows by Column C (index 2).
        # We put non-empty URLs first sorted A-Z, followed by empty URLs.
        def sort_key(row):
            url = row[2].strip() if len(row) > 2 else ""
            # If empty, return (1, "") so it comes after non-empty (0, url)
            return (0, url.lower()) if url else (1, "")

        sorted_data = sorted(data_rows, key=sort_key)
        
        # Count rows with invitation URLs
        filled_count = sum(1 for r in sorted_data if len(r) > 2 and r[2].strip())

        with open(NEW_CSV_PATH, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            writer.writerows(sorted_data)

        fcntl.flock(lock_f, fcntl.LOCK_UN)

    print(f"✅ NEW.csv sorted by Column C (A-Z)!")
    print(f"📊 Total rows with Invite URLs in Column C (placed at the top): {filled_count}")

def sort_excel_by_invite_url():
    if not EXCEL_PATH.exists():
        return

    excel_lock_path = EXCEL_PATH.with_suffix(".xlsx.lock")
    with open(excel_lock_path, "w") as lock_f:
        fcntl.flock(lock_f, fcntl.LOCK_EX)
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Talabalar"]

        # Read all rows
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return

        header = rows[0]
        data_rows = rows[1:]

        # Column F is Invite Link (index 5 in 0-indexed list)
        def sort_key(row):
            url = str(row[5] or "").strip() if len(row) > 5 else ""
            return (0, url.lower()) if url else (1, "")

        sorted_data = sorted(data_rows, key=sort_key)

        # Clear existing sheet data
        ws.delete_rows(1, ws.max_row)

        # Write header and sorted rows
        ws.append(list(header))
        for row in sorted_data:
            ws.append(list(row))

        wb.save(EXCEL_PATH)
        fcntl.flock(lock_f, fcntl.LOCK_UN)

    print(f"✅ Names.xlsx sorted by Column F (Invite Link A-Z)!")

if __name__ == "__main__":
    sort_new_csv_by_column_c()
    sort_excel_by_invite_url()
