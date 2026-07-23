import os
import re
import json
import imaplib
import email
import requests
import openpyxl
import fcntl
import time
import threading
import socket
import concurrent.futures
from pathlib import Path

# Set global socket timeout to prevent any thread from hanging indefinitely
socket.setdefaulttimeout(15)

# Config
EXCEL_PATH = Path("Names.xlsx")
DB_PATH = Path("Names_db.json")
GMAIL_USER = "qwertyuioplkjhgfdsazxcvbnmhrh@gmail.com"
GMAIL_APP_PASS = "feykxtyuitmnjevn"
CONCURRENCY = 5
excel_lock_path = Path("excel.lock")

# We only search emails received since we started registrations (July 15, 2026)
SINCE_DATE = "15-Jul-2026"

def resolve_redirect(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    try:
        response = requests.head(url, headers=headers, allow_redirects=True, timeout=10)
        return response.url
    except Exception:
        try:
            response = requests.get(url, headers=headers, allow_redirects=True, timeout=10)
            return response.url
        except Exception:
            return url

def normalize_email(email_addr):
    if not email_addr or "@" not in email_addr:
        return ""
    return email_addr.strip().lower()

def clean_name_words(name_str):
    if not name_str:
        return []
    cleaned = name_str.lower().replace("‘", "").replace("’", "").replace("'", "").replace("`", "")
    words = [re.sub(r'[^a-z0-9]', '', w) for w in cleaned.split()]
    return [w for w in words if len(w) > 2]

def split_list(lst, n_chunks):
    k, m = divmod(len(lst), n_chunks)
    return [lst[i*k+min(i, m):(i+1)*k+min(i+1, m)] for i in range(n_chunks)]

def fetch_worker(worker_id, folder, msg_ids_chunk, email_to_student, name_to_student):
    results = {}
    if not msg_ids_chunk:
        return results

    # Reverse chunk to prioritize the absolute newest emails first
    msg_ids_chunk = list(reversed(msg_ids_chunk))

    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com", timeout=15)
        mail.login(GMAIL_USER, GMAIL_APP_PASS)
        status, _ = mail.select(f'"{folder}"', readonly=True)
        if status != "OK":
            status, _ = mail.select(folder, readonly=True)
            if status != "OK":
                mail.logout()
                return results

        chunk_size = 100
        total_batches = (len(msg_ids_chunk) + chunk_size - 1) // chunk_size
        for idx, i in enumerate(range(0, len(msg_ids_chunk), chunk_size), start=1):
            batch = msg_ids_chunk[i:i+chunk_size]
            batch_str = ",".join(x.decode("utf-8") for x in batch)
            
            print(f"  ✉️ [Worker {worker_id}] Fetching batch {idx}/{total_batches} ({len(batch)} headers)...", flush=True)
            # Fetch only TO and SUBJECT headers (extremely fast bulk operation)
            status, fetch_data = mail.fetch(batch_str, "(BODY[HEADER.FIELDS (TO SUBJECT)])")
            if status != "OK":
                continue

            valid_msgs = []
            for item in fetch_data:
                if not isinstance(item, tuple):
                    continue
                
                header_part = item[1]
                header_info = item[0].decode('utf-8', errors='ignore')
                msg_id_match = re.search(r'^\d+', header_info)
                if not msg_id_match:
                    continue
                msg_id = msg_id_match.group(0)
                
                msg = email.message_from_bytes(header_part)
                to_header = msg.get("To") or ""
                to_emails = re.findall(r'[\w\.-]+@[\w\.-]+', to_header)
                to_email = to_emails[0].strip().lower() if to_emails else ""
                norm_to = normalize_email(to_email)
                
                subject = msg.get("Subject") or ""
                
                # Match by email or try fallback name match if subject looks like a Coursera invitation
                matched_student = None
                if norm_to in email_to_student:
                    matched_student = email_to_student[norm_to]
                elif "invited" in subject.lower() or "invitation" in subject.lower():
                    # Deep match using name words
                    subject_words = clean_name_words(subject)
                    for key_words, student in name_to_student.items():
                        if all(w in subject_words for w in key_words):
                            matched_student = student
                            break

                if matched_student:
                    valid_msgs.append((msg_id, matched_student))

            # Fetch bodies only for matched messages
            for msg_id, student in valid_msgs:
                try:
                    status, body_data = mail.fetch(msg_id, "(BODY[TEXT])")
                    if status != "OK":
                        continue
                    
                    body_text = ""
                    for part in body_data:
                        if isinstance(part, tuple):
                            try:
                                body_text += part[1].decode("utf-8", errors="ignore")
                            except Exception:
                                body_text += str(part[1])

                    # Extract Coursera program invite URL
                    links = re.findall(r'https://[^\s"<>]+coursera\.org/[^\s"<>]+', body_text)
                    invite_link = ""
                    for l in links:
                        if "invitationToken=" in l or "program" in l:
                            invite_link = l.split("?")[0] + "?" + l.split("?")[1].split()[0]
                            invite_link = invite_link.rstrip('>').rstrip('"').rstrip(')').rstrip('.')
                            break

                    if invite_link:
                        # Resolve redirect
                        resolved_url = resolve_redirect(invite_link)
                        print(f"  ✨ [Worker {worker_id}] Found invite for Row {student['row']}: {student['name']} -> {resolved_url}", flush=True)
                        results[student["row"]] = {
                            "invite_url": resolved_url,
                            "email": student["email"],
                            "passport": student["passport"]
                        }
                except Exception as e:
                    print(f"⚠️ [Worker {worker_id}] Error processing message {msg_id}: {e}", flush=True)

        mail.logout()
    except Exception as e:
        print(f"⚠️ [Worker {worker_id}] Error: {e}", flush=True)

    return results

def main():
    print("📖 Loading target list from Names.xlsx and Names_db.json...")
    
    # Load JSON DB
    db_by_passport = {}
    if DB_PATH.exists():
        try:
            db_data = json.loads(DB_PATH.read_text())
            db_by_passport = {item["document"]: item for item in db_data if item.get("document")}
        except Exception as e:
            print(f"⚠️ Could not load database JSON: {e}")

    # Load Excel sheet
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Talabalar"]

    email_to_student = {}
    name_to_student = {}

    for r_idx in range(3, 5200):
        passport = str(ws.cell(row=r_idx, column=3).value or "").strip()
        if not passport:
            continue

        existing_link = ws.cell(row=r_idx, column=6).value
        has_link = existing_link and "coursera.org" in str(existing_link)

        email_addr = str(ws.cell(row=r_idx, column=7).value or "").strip().lower()
        if not email_addr and passport in db_by_passport:
            email_addr = str(db_by_passport[passport].get("email") or "").strip().lower()

        student_name = str(ws.cell(row=r_idx, column=1).value or "").strip()

        # Target students: registered email but no invite link yet
        if email_addr and not has_link:
            student_record = {
                "row": r_idx,
                "passport": passport,
                "name": student_name,
                "email": email_addr,
                "norm_email": normalize_email(email_addr),
                "name_words": clean_name_words(student_name)
            }
            email_to_student[student_record["norm_email"]] = student_record
            
            if len(student_record["name_words"]) >= 2:
                name_key = tuple(sorted(student_record["name_words"][:2]))
                name_to_student[name_key] = student_record

    print(f"Loaded {len(email_to_student)} students with registered emails but missing invite links.", flush=True)
    if not email_to_student:
        print("All target students already have invite links. Exiting.")
        return

    folders = ["INBOX", "[Gmail]/Spam", "[Gmail]/All Mail"]
    updates_found = {}
    save_lock = threading.Lock()

    def incremental_save():
        """Incremental save function to write updates immediately to disk."""
        with save_lock:
            if not updates_found:
                return
            print(f"\n💾 Incremental save: writing {len(updates_found)} links to disk...", flush=True)

            # Update Names.xlsx
            try:
                with open(excel_lock_path, "w") as lock_f:
                    fcntl.flock(lock_f, fcntl.LOCK_EX)
                    wb_w = openpyxl.load_workbook(EXCEL_PATH)
                    ws_w = wb_w["Talabalar"]
                    for row_idx, data in updates_found.items():
                        ws_w.cell(row=row_idx, column=6, value=data["invite_url"])
                        ws_w.cell(row=row_idx, column=7, value=data["email"])
                    wb_w.save(EXCEL_PATH)
                    fcntl.flock(lock_f, fcntl.LOCK_UN)
                print(f"  ✅ Names.xlsx updated ({len(updates_found)} rows).", flush=True)
            except Exception as err:
                print(f"  ❌ Failed to write Names.xlsx: {err}", flush=True)

            # Update JSON DB
            if DB_PATH.exists():
                lock_path = DB_PATH.with_suffix(".json.lock")
                try:
                    with open(lock_path, "w") as lock_f2:
                        fcntl.flock(lock_f2, fcntl.LOCK_EX)
                        db = json.loads(DB_PATH.read_text())
                        passport_updates = {data["passport"]: data["invite_url"] for data in updates_found.values()}
                        for item in db:
                            if item.get("document") in passport_updates:
                                item["invite_url"] = passport_updates[item["document"]]
                                item["activated"] = True
                        DB_PATH.write_text(json.dumps(db, ensure_ascii=False, indent=2))
                        fcntl.flock(lock_f2, fcntl.LOCK_UN)
                    print(f"  ✅ Names_db.json updated.", flush=True)
                except Exception as err:
                    print(f"  ❌ Failed to write Names_db.json: {err}", flush=True)

    executor = concurrent.futures.ThreadPoolExecutor(max_workers=CONCURRENCY)

    for folder in folders:
        try:
            print(f"\n📂 Scanning folder: {folder}...", flush=True)
            mail_temp = imaplib.IMAP4_SSL("imap.gmail.com", timeout=15)
            mail_temp.login(GMAIL_USER, GMAIL_APP_PASS)
            status, _ = mail_temp.select(f'"{folder}"', readonly=True)
            if status != "OK":
                status, _ = mail_temp.select(folder, readonly=True)
                if status != "OK":
                    mail_temp.logout()
                    continue

            # Optimize: only retrieve messages since the start date
            status, data = mail_temp.search(None, f'(SINCE "{SINCE_DATE}")')
            if status != "OK" or not data[0]:
                mail_temp.logout()
                print(f"  ↳ No messages since {SINCE_DATE}.", flush=True)
                continue

            msg_ids = data[0].split()
            mail_temp.logout()

            print(f"Found {len(msg_ids)} messages since {SINCE_DATE} in {folder}. Running with {CONCURRENCY} threads...", flush=True)

            chunks = split_list(msg_ids, CONCURRENCY)
            
            futures_dict = {}
            for w_id, chunk in enumerate(chunks, start=1):
                if not chunk:
                    continue
                f = executor.submit(fetch_worker, w_id, folder, chunk, email_to_student, name_to_student)
                futures_dict[f] = w_id
            
            # Wait until all threads complete this folder
            done, not_done = concurrent.futures.wait(futures_dict.keys(), timeout=None)
            
            for future in done:
                try:
                    res = future.result()
                    if res:
                        updates_found.update(res)
                except Exception as err:
                    w_id = futures_dict[future]
                    print(f"⚠️ [Worker {w_id}] raised an exception: {err}", flush=True)

            if not_done:
                for future in not_done:
                    w_id = futures_dict[future]
                    print(f"⚠️ [Worker {w_id}] was slow/stuck in {folder} and bypassed.", flush=True)

            # Save incrementally after completing/timing out folder
            incremental_save()

        except Exception as e:
            print(f"⚠️ Error scanning folder {folder}: {e}", flush=True)

    # Shut down the executor cleanly after all folders are processed
    executor.shutdown(wait=False)

    # Final forced save of all results
    incremental_save()
    print(f"\nDone! Successfully resolved and synchronized {len(updates_found)} invitation links.", flush=True)

if __name__ == "__main__":
    main()
