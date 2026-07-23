#!/usr/bin/env python3
import os
import sys
import re
import csv
import json
import time
import imaplib
import email
import random
import subprocess
import urllib.parse
from pathlib import Path

# Paths
EXCEL_PATH = Path("Names.xlsx")
CSV_PATH = Path("students.csv")
RESULTS_PATH = Path("results.json")
PROCESSED_PATH = Path("processed_invites.json")

# Hardcoded GMAIL Credentials (matching registration_automation.py)
GMAIL_USER = "qwertyuioplkjhgfdsazxcvbnmhrh@gmail.com"
GMAIL_APP_PASS = "feykxtyuitmnjevn".replace(" ", "")

# Load Openpyxl
try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Run: pip install openpyxl")
    sys.exit(1)


class GmailIMAPClient:
    def __init__(self, username, app_password):
        self.username = username
        self.app_password = app_password
        self.mail = None

    def connect(self):
        try:
            if self.mail:
                self.mail.logout()
        except Exception:
            pass
        self.mail = imaplib.IMAP4_SSL("imap.gmail.com")
        self.mail.login(self.username, self.app_password)

    def ensure_connected(self):
        if not self.mail:
            self.connect()
            return
        try:
            status, _ = self.mail.noop()
            if status != "OK":
                self.connect()
        except Exception:
            self.connect()

    def check_for_coursera_invites(self, processed_ids):
        self.ensure_connected()
        folders = ["INBOX", "[Gmail]/Spam", "[Gmail]/All Mail"]
        invites = []

        for folder in folders:
            try:
                self.ensure_connected()
                status, _ = self.mail.select(f'"{folder}"', readonly=True)
                if status != "OK":
                    status, _ = self.mail.select(folder, readonly=True)
                    if status != "OK":
                        continue

                # Search for emails from coursera or containing coursera in the subject/body
                status, data = self.mail.search(None, 'OR (FROM "coursera.org") (SUBJECT "invite")')
                if status != "OK":
                    continue

                msg_ids = data[0].split()
                if not msg_ids:
                    continue

                # Process the latest 20 messages in this folder to avoid overhead
                for msg_id in reversed(msg_ids[-20:]):
                    status, msg_data = self.mail.fetch(msg_id, "(RFC822)")
                    if status != "OK":
                        continue

                    raw_email = msg_data[0][1]
                    msg = email.message_from_bytes(raw_email)

                    msg_id_header = msg.get("Message-ID") or ""
                    msg_key = msg_id_header.strip() if msg_id_header else f"{folder}:{msg_id.decode('utf-8')}"
                    if msg_key in processed_ids:
                        continue

                    to_header = msg.get("To") or ""
                    subject = msg.get("Subject") or ""
                    
                    body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            content_type = part.get_content_type()
                            content_disp = str(part.get("Content-Disposition"))
                            if content_type == "text/html" and "attachment" not in content_disp:
                                body = part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", errors="ignore")
                                break
                            elif content_type == "text/plain" and "attachment" not in content_disp:
                                body = part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", errors="ignore")
                    else:
                        body = msg.get_payload(decode=True).decode(msg.get_content_charset() or "utf-8", errors="ignore")

                    # Extract Coursera link
                    all_matches = re.findall(r'https?://[^\s\"\'\>]+', body)
                    invite_url = None
                    
                    for raw_url in all_matches:
                        clean_url = "".join(raw_url.split()).replace("&amp;", "&").rstrip(".,;!?'\"<>#")
                        
                        # If the link already contains invitationToken directly in the email body
                        if "coursera.org" in clean_url and "invitationToken=" in clean_url:
                            invite_url = clean_url
                            break
                            
                    if not invite_url:
                        # If not found directly, resolve all link.coursera.org/f/ links and take the one with invitationToken
                        for raw_url in all_matches:
                            clean_url = "".join(raw_url.split()).replace("&amp;", "&").rstrip(".,;!?'\"<>#")
                            if "link.coursera.org/f/" in clean_url:
                                try:
                                    resolved = resolve_redirect(clean_url)
                                    if "invitationToken=" in resolved:
                                        invite_url = resolved
                                        break
                                except Exception:
                                    pass

                    if invite_url:
                        emails_found = re.findall(r'[\w\.-]+@[\w\.-]+', to_header)
                        to_email = emails_found[0] if emails_found else ""

                        invites.append({
                            "msg_key": msg_key,
                            "to_email": to_email,
                            "url": invite_url,
                            "subject": subject
                        })
            except Exception as e:
                print(f"      [IMAP Check Warning in folder {folder}]: {e}")

        return invites

    def disconnect(self):
        try:
            if self.mail:
                self.mail.logout()
        except Exception:
            pass


def normalize_email(email_str):
    email_str = email_str.strip().lower()
    if "@gmail.com" in email_str:
        username, domain = email_str.split("@")
        return username.replace(".", "") + "@" + domain
    return email_str


def load_processed_invites():
    if PROCESSED_PATH.exists():
        try:
            return set(json.loads(PROCESSED_PATH.read_text()))
        except Exception:
            return set()
    return set()


def save_processed_invite(msg_key):
    processed = load_processed_invites()
    processed.add(msg_key)
    PROCESSED_PATH.write_text(json.dumps(list(processed), indent=2))


def lookup_student_in_results(to_email):
    if not RESULTS_PATH.exists():
        return None
    try:
        results = json.loads(RESULTS_PATH.read_text())
        norm_to = normalize_email(to_email)
        for entry in results:
            if "email" in entry and entry["email"]:
                if normalize_email(entry["email"]) == norm_to:
                    return entry
    except Exception as e:
        print(f"Error reading results.json: {e}")
    return None


import requests

def resolve_redirect(url):
    try:
        response = requests.head(url, allow_redirects=True, timeout=10)
        return response.url
    except Exception:
        try:
            response = requests.get(url, allow_redirects=True, timeout=10)
            return response.url
        except Exception:
            return url

def get_student_details_from_excel(passport_num):
    if not EXCEL_PATH.exists():
        print(f"Excel file {EXCEL_PATH} not found.")
        return None
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        sheet = wb["Talabalar"]
        
        found_row = -1
        for r in range(3, sheet.max_row + 1):
            cell_val = str(sheet.cell(r, 2).value or "").strip()
            if cell_val == passport_num:
                found_row = r
                break
                
        if found_row == -1:
            print(f"⚠️ Warning: Passport number {passport_num} not found in Excel roster.")
            return None
            
        full_name = str(sheet.cell(found_row, 1).value or "").strip()
        passport = str(sheet.cell(found_row, 2).value or "").strip()
        jshshir = str(sheet.cell(found_row, 3).value or "").strip()
        
        if not full_name or not passport:
            return None
            
        parts = full_name.split()
        first_name = parts[0] if parts else ""
        last_name = " ".join(parts[1:]) if len(parts) > 1 else ""
        
        return {
            "student_id": jshshir if jshshir else passport,
            "first_name": first_name,
            "last_name": last_name,
            "password": passport,
            "row": found_row
        }
    except Exception as e:
        print(f"Error searching Excel for passport {passport_num}: {e}")
    return None


def update_or_add_student_in_csv(details, email_addr):
    students = []
    headers = ["student_id", "first_name", "last_name", "email", "certificate_url", "password"]
    
    if CSV_PATH.exists():
        try:
            with open(CSV_PATH, "r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or headers
                students = list(reader)
        except Exception as e:
            print(f"Error reading students.csv: {e}")
            
    # Check if student exists
    found_idx = -1
    for idx, s in enumerate(students):
        if s.get("student_id") == details["student_id"] or s.get("password") == details["password"]:
            found_idx = idx
            break
            
    student_row = {
        "student_id": details["student_id"],
        "first_name": details["first_name"],
        "last_name": details["last_name"],
        "email": email_addr,
        "certificate_url": s.get("certificate_url", "") if found_idx != -1 else "",
        "password": details["password"]
    }
    
    # Fill remaining headers if any
    for h in headers:
        if h not in student_row:
            student_row[h] = ""
            
    if found_idx != -1:
        students[found_idx] = student_row
        student_index = found_idx + 1 # 1-based index (index 0 is row 2, which is student 1)
        print(f"Updated existing student {details['first_name']} {details['last_name']} in students.csv at index {student_index}.")
    else:
        students.append(student_row)
        student_index = len(students) # 1-based index of newly appended student
        print(f"Appended new student {details['first_name']} {details['last_name']} to students.csv at index {student_index}.")
        
    try:
        with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(students)
    except Exception as e:
        print(f"Error writing students.csv: {e}")
        
    return student_index


def check_if_cert_obtained(details):
    if not CSV_PATH.exists():
        return None
    try:
        with open(CSV_PATH, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for s in reader:
                if s.get("student_id") == details["student_id"] or s.get("password") == details["password"]:
                    cert = s.get("certificate_url", "").strip()
                    return cert if cert else None
    except Exception:
        pass
    return None


def run_replay_for_student(index, full_name, invite_url):
    print(f"\n🚀 Starting automated replay subprocess for student: {full_name}")
    print(f"🔗 URL: {invite_url}")
    print(f"Index in CSV: {index}\n" + "-"*50)
    
    env = os.environ.copy()
    env["START"] = str(index)
    env["END"] = str(index)
    env["CONCURRENCY"] = "1"
    env["COURSE_URL"] = invite_url
    env["MODE"] = "auto"
    
    cmd = ["node", "coursera_manual_runner.js"]
    try:
        process = subprocess.Popen(cmd, env=env, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
        for line in process.stdout:
            sys.stdout.write(line)
            sys.stdout.flush()
        process.wait()
        print("-"*50 + f"\nReplay subprocess exited with code {process.returncode}.")
        return process.returncode == 0
    except Exception as e:
        print(f"❌ Subprocess failed to execute: {e}")
        return False


def main():
    print("=" * 60)
    print("Coursera Invitation Email Daemon Started")
    print(f"Monitoring Gmail: {GMAIL_USER}")
    print("Waiting for new Coursera invitation links...")
    print("=" * 60)

    imap_client = GmailIMAPClient(GMAIL_USER, GMAIL_APP_PASS)
    
    try:
        imap_client.connect()
        print("✅ Connected to Gmail IMAP successfully.")
    except Exception as e:
        print(f"❌ Connection failed: {e}")
        return

    processed_ids = load_processed_invites()
    print(f"Loaded {len(processed_ids)} previously processed email IDs.")

    try:
        while True:
            try:
                print(f"\n[{time.strftime('%Y-%m-%d %H:%M:%S')}] Polling for new invites...")
                invites = imap_client.check_for_coursera_invites(processed_ids)
                
                if invites:
                    print(f"Found {len(invites)} new potential invite emails.")
                    for invite in invites:
                        to_email = invite["to_email"]
                        invite_url = invite["url"]
                        if "link.coursera.org" in invite_url or "google.com/url" in invite_url:
                            try:
                                print(f"  Resolving tracking link: {invite_url} ...")
                                invite_url = resolve_redirect(invite_url)
                                print(f"  Resolved to: {invite_url}")
                            except Exception as res_err:
                                print(f"  ⚠️ Warning resolving redirect: {res_err}")

                        msg_key = invite["msg_key"]
                        
                        print(f"\nProcessing invite email:")
                        print(f"  To:      {to_email}")
                        print(f"  Subject: {invite['subject']}")
                        print(f"  Link:    {invite_url}")

                        # Match student
                        reg_info = lookup_student_in_results(to_email)
                        if not reg_info:
                            print(f"⚠️ Warning: Could not find matching registration in results.json for {to_email}. Skipping.")
                            processed_ids.add(msg_key)
                            save_processed_invite(msg_key)
                            continue

                        passport_num = reg_info.get("document")
                        if not passport_num:
                            print(f"⚠️ Warning: Matched entry has no document/passport. Skipping.")
                            processed_ids.add(msg_key)
                            save_processed_invite(msg_key)
                            continue

                        # Read name from Excel by searching for passport
                        details = get_student_details_from_excel(passport_num)
                        if not details:
                            print(f"⚠️ Warning: Failed to load student details from Names.xlsx for passport {passport_num}. Skipping.")
                            processed_ids.add(msg_key)
                            save_processed_invite(msg_key)
                            continue

                        full_name = f"{details['first_name']} {details['last_name']}"
                        print(f"👤 Matched Student: {full_name} (Excel Row {details['row']}, Passport: {details['password']})")

                        # Update students.csv
                        student_index = update_or_add_student_in_csv(details, to_email)

                        # Run replay
                        success = run_replay_for_student(student_index, full_name, invite_url)
                        
                        # Verify results
                        cert_url = check_if_cert_obtained(details)
                        if cert_url:
                            print(f"\n🎉 SUCCESS! Certificate obtained for {full_name}!")
                            print(f"📜 Link: {cert_url}")
                            try:
                                if RESULTS_PATH.exists():
                                    results = json.loads(RESULTS_PATH.read_text())
                                    for entry in results:
                                        if entry.get("document") == details["password"]:
                                            entry["certificate_url"] = cert_url
                                            entry["activated"] = True
                                            break
                                    RESULTS_PATH.write_text(json.dumps(results, ensure_ascii=False, indent=2))
                                    print(f"💾 Saved certificate URL to results.json for Doc={details['password']}.")
                            except Exception as json_err:
                                print(f"⚠️ Warning: Could not write certificate to results.json: {json_err}")
                        else:
                            print(f"\n❌ Failure: Replay finished but no certificate URL saved for {full_name}.")

                        # Mark as processed
                        processed_ids.add(msg_key)
                        save_processed_invite(msg_key)
                        
                else:
                    print("No new invite emails found.")

            except Exception as e:
                print(f"Error during polling iteration: {e}")
                
            time.sleep(30)

    except KeyboardInterrupt:
        print("\nStopping daemon...")
    finally:
        imap_client.disconnect()
        print("Disconnected from Gmail IMAP.")


if __name__ == "__main__":
    main()
