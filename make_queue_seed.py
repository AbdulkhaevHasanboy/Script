#!/usr/bin/env python3
"""Export names.xlsx -> queue_seed.csv (student_id, full_name).

This produces the two columns you import into the coordinator Google Sheet to
seed the distributed work queue. Reads the same "Talabalar" sheet (from row 3)
that the Node runner uses, so the queue lines up exactly with the student list.

Usage:
    python3 make_queue_seed.py            # reads names.xlsx -> queue_seed.csv
    python3 make_queue_seed.py in.xlsx out.csv
"""

import csv
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install it with:  pip install openpyxl")

SRC = sys.argv[1] if len(sys.argv) > 1 else "names.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "queue_seed.csv"
SHEET = "Talabalar"


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Sheet '{SHEET}' not found in {SRC}. Tabs: {wb.sheetnames}")
    sheet = wb[SHEET]

    rows = []
    for r in range(3, sheet.max_row + 1):
        sid = sheet.cell(r, 1).value
        name = sheet.cell(r, 2).value
        if sid and name:
            rows.append([str(sid).strip(), str(name).strip()])

    with open(OUT, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["student_id", "full_name"])
        w.writerows(rows)

    print(f"Wrote {len(rows)} students to {OUT}")
    print("Next: open your coordinator Google Sheet, select the 'Queue' tab, "
          "click cell A1, then File > Import > Upload this CSV > "
          "'Replace data at selected cell'. Then run initQueue() once in the "
          "Apps Script editor.")


if __name__ == "__main__":
    main()
