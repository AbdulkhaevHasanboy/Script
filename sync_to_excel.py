import json
import openpyxl
import fcntl
from pathlib import Path

# Config
EXCEL_PATH = Path("Names.xlsx")
DB_PATH = Path("Names_db.json")

def main():
    if not DB_PATH.exists():
        print("Error: Names_db.json not found.")
        return
    if not EXCEL_PATH.exists():
        print("Error: Names.xlsx not found.")
        return

    # Load Names_db.json with process lock
    lock_path = DB_PATH.with_suffix(".json.lock")
    try:
        with open(lock_path, "w") as lock_f:
            fcntl.flock(lock_f, fcntl.LOCK_EX)
            db = json.loads(DB_PATH.read_text())
    except Exception as e:
        print(f"Error reading DB: {e}")
        return

    print(f"Loaded {len(db)} students from database.")

    # Excel save process lock
    excel_lock_path = EXCEL_PATH.with_suffix(".xlsx.lock")
    try:
        with open(excel_lock_path, "w") as lock_f:
            fcntl.flock(lock_f, fcntl.LOCK_EX)
            
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb["Talabalar"]
            
            # Ensure headers in F and G
            ws.cell(row=1, column=6, value="Invite Link")
            ws.cell(row=1, column=7, value="Email")

            # Map passport to sheet row
            passport_to_row = {}
            for r in range(2, ws.max_row + 1):
                pass_val = str(ws.cell(row=r, column=2).value or "").strip()
                if pass_val:
                    passport_to_row[pass_val] = r

            # Write emails and links from DB
            written_emails = 0
            written_links = 0
            
            for entry in db:
                passport = entry.get("document")
                email_addr = entry.get("email")
                invite_url = entry.get("invite_url")

                row_idx = passport_to_row.get(passport)
                if row_idx:
                    if email_addr:
                        ws.cell(row=row_idx, column=7, value=email_addr)
                        written_emails += 1
                    if invite_url and "coursera.org" in invite_url:
                        ws.cell(row=row_idx, column=6, value=invite_url)
                        written_links += 1

            wb.save(EXCEL_PATH)
            print(f"Success! Synchronized {written_emails} emails and {written_links} Coursera links to Names.xlsx.")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

if __name__ == "__main__":
    main()
