import json
from openpyxl import load_workbook
from pathlib import Path

# Load Names.xlsx
wb = load_workbook("Names.xlsx", data_only=True)
ws = wb["Talabalar"]

# Load existing results.json if it exists
results_map = {}
if Path("results.json").exists():
    try:
        res_data = json.loads(Path("results.json").read_text())
        for r in res_data:
            if r.get("document"):
                results_map[r["document"]] = r
    except Exception as e:
        print(f"Warning: could not parse results.json: {e}")

headers = [str(h).strip() if h is not None else "" for h in next(ws.iter_rows(max_row=1, values_only=True))]
name_idx = headers.index("Full Name of student")
doc_idx = headers.index("Pasport raqami")
dob_idx = headers.index("Tug‘ilgan sana")
phone_idx = headers.index("Telefon")

db_records = []
for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if not row or row[doc_idx] is None:
        continue
    
    document = str(row[doc_idx]).strip()
    dob_val = row[dob_idx]
    if hasattr(dob_val, "strftime"):
        dob = dob_val.strftime("%Y-%m-%d")
    else:
        dob = str(dob_val).strip()
        if len(dob) >= 10 and dob[4] == '-' and dob[7] == '-':
            dob = dob[:10]
            
    phone_raw = str(row[phone_idx]).strip() if row[phone_idx] is not None else ""
    phone_clean = "".join(c for c in phone_raw if c.isdigit())
    if phone_clean.startswith("998"):
        phone = f"+{phone_clean}"
    elif len(phone_clean) == 9:
        phone = f"+998{phone_clean}"
    else:
        phone = f"+{phone_clean}"

    name = str(row[name_idx]).strip()
    
    excel_invite = str(row[5] or "").strip() if len(row) > 5 else ""
    excel_email = str(row[6] or "").strip() if len(row) > 6 else ""

    res_entry = results_map.get(document, {})
    email = res_entry.get("email") or excel_email or None
    invite_url = res_entry.get("invite_url") or res_entry.get("activation_url") or excel_invite or None
    if invite_url and "coursera.org" not in invite_url:
        invite_url = None
    activated = res_entry.get("activated", False) or (True if email else False)

    db_records.append({
        "row": r_idx,
        "name": name,
        "document": document,
        "dob": dob,
        "phone": phone,
        "email": email,
        "invite_url": invite_url,
        "activated": activated
    })

# Write to Names_db.json
Path("Names_db.json").write_text(json.dumps(db_records, ensure_ascii=False, indent=2))
print(f"Success! Converted {len(db_records)} records to Names_db.json.")
