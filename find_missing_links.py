import json
import re
import imaplib
import email
import openpyxl
import fcntl
import time
import threading
import concurrent.futures
from pathlib import Path

EXCEL_PATH = Path("Names.xlsx")
DB_PATH = Path("Names_db.json")
GMAIL_USER = "qwertyuioplkjhgfdsazxcvbnmhrh@gmail.com"
GMAIL_APP_PASS = "feykxtyuitmnjevn"
CONCURRENCY = 10
excel_lock_path = Path("excel.lock")

COURSERA_LINK_RE = re.compile(
    r'https://www\.coursera\.org/programs/[^\s"<>]+invitationToken=[^\s"<>]+'
)

def extract_link_from_body(raw_body: bytes) -> str:
    try:
        text = raw_body.decode("utf-8", errors="replace")
    except Exception:
        text = str(raw_body)

    text = re.sub(r'=\r?\n', '', text)
    text = re.sub(r'=([0-9A-Fa-f]{2})', lambda m: chr(int(m.group(1), 16)), text)
    text = re.sub(r'\s+', ' ', text)

    match = COURSERA_LINK_RE.search(text)
    if match:
        return match.group(0).rstrip('.')
    return ""

def scan_chunk_in_folder(worker_id, folder, student_chunk, results, lock):
    """Worker function to search a chunk of students inside a single pre-selected folder."""
    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(GMAIL_USER, GMAIL_APP_PASS)
        status, _ = mail.select(f'"{folder}"', readonly=True)
        if status != "OK":
            status, _ = mail.select(folder, readonly=True)
            if status != "OK":
                mail.logout()
                return
    except Exception as e:
        print(f"  ❌ [Worker {worker_id}] Connection/Select failed: {e}")
        return

    for idx, student in enumerate(student_chunk):
        email_addr = student["email"]
        row = student["row"]
        name = student["name"]
        
        try:
            status, data = mail.search(None, f'(TO "{email_addr}")')
            if status != "OK" or not data[0]:
                continue

            msg_ids = data[0].split()
            if not msg_ids:
                continue

            # Fetch the newest matching message
            newest_id = msg_ids[-1]
            status, fetch_data = mail.fetch(newest_id, "(RFC822)")
            if status != "OK":
                continue

            for part in fetch_data:
                if isinstance(part, tuple):
                    link = extract_link_from_body(part[1])
                    if link:
                        print(f"  ✨ [Worker {worker_id}] Found link for Row {row}: {name}")
                        with lock:
                            results[row] = {
                                "invite_url": link,
                                "email": email_addr,
                                "passport": student.get("document", "")
                            }
                        break
        except Exception as e:
            err = str(e)
            if "Too many simultaneous" in err:
                time.sleep(2)
            # ignore other search-specific errors and continue

    try:
        mail.logout()
    except Exception:
        pass

def save_to_disk(results):
    if not results:
        return

    print(f"\n💾 Writing {len(results)} new links to disk...")

    # Update Excel
    try:
        with open(excel_lock_path, "w") as lock_f:
            fcntl.flock(lock_f, fcntl.LOCK_EX)
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb["Talabalar"]
            for row_idx, data in results.items():
                ws.cell(row=row_idx, column=6, value=data["invite_url"])
                ws.cell(row=row_idx, column=7, value=data["email"])
            wb.save(EXCEL_PATH)
            fcntl.flock(lock_f, fcntl.LOCK_UN)
        print("  ✅ Names.xlsx updated successfully.")
    except Exception as err:
        print(f"  ❌ Failed to write to Names.xlsx: {err}")

    # Update DB
    if DB_PATH.exists():
        lock_path = DB_PATH.with_suffix(".json.lock")
        try:
            with open(lock_path, "w") as lock_f:
                fcntl.flock(lock_f, fcntl.LOCK_EX)
                db = json.loads(DB_PATH.read_text())
                passport_map = {v["passport"]: v["invite_url"] for v in results.values() if v.get("passport")}
                for item in db:
                    if item.get("document") in passport_map:
                        item["invite_url"] = passport_map[item["document"]]
                        item["activated"] = True
                DB_PATH.write_text(json.dumps(db, ensure_ascii=False, indent=2))
                fcntl.flock(lock_f, fcntl.LOCK_UN)
            print("  ✅ Names_db.json updated successfully.")
        except Exception as err:
            print(f"  ❌ Failed to write to Names_db.json: {err}")

def main():
    print("🔍 Loading missing students from Names_db.json...")
    db = json.loads(DB_PATH.read_text())

    missing = [
        r for r in db
        if r.get("email")
        and r.get("row")
        and not (r.get("invite_url") and "coursera.org" in str(r.get("invite_url", "")))
    ]
    
    print(f"  → {len(missing)} students currently missing invite links.")
    if not missing:
        print("No missing links to search! Exiting.")
        return

    folders = ["INBOX", "[Gmail]/Spam", "[Gmail]/All Mail"]
    total_found = 0

    for folder in folders:
        if not missing:
            break

        print(f"\n📂 Scanning folder: {folder} for {len(missing)} missing students...")
        
        # Split currently missing students into chunks for the 10 workers
        chunk_size = max(1, len(missing) // CONCURRENCY)
        chunks = [missing[i:i+chunk_size] for i in range(0, len(missing), chunk_size)]
        
        folder_results = {}
        lock = threading.Lock()
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=CONCURRENCY) as executor:
            futures = []
            for w_id, chunk in enumerate(chunks, start=1):
                if chunk:
                    f = executor.submit(scan_chunk_in_folder, w_id, folder, chunk, folder_results, lock)
                    futures.append(f)
            
            concurrent.futures.wait(futures)

        if folder_results:
            print(f"  ↳ Found {len(folder_results)} new links in {folder}!")
            save_to_disk(folder_results)
            total_found += len(folder_results)
            
            # Remove found students from the missing list
            found_rows = set(folder_results.keys())
            missing = [student for student in missing if student["row"] not in found_rows]
        else:
            print(f"  ↳ No links found in {folder}.")

    print(f"\n🏁 Finished search. Total new links saved: {total_found}")

if __name__ == "__main__":
    main()
