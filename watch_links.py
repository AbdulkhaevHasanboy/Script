import os
import re
import json
import csv
import time
import imaplib
import email
import requests
import openpyxl
from pathlib import Path

# Config
DB_PATH = Path("Names_db.json")
PROCESSED_PATH = Path("processed_invites.json")
GMAIL_USER = "qwertyuioplkjhgfdsazxcvbnmhrh@gmail.com"
GMAIL_APP_PASS = "feykxtyuitmnjevn"

def resolve_redirect(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    try:
        response = requests.head(url, headers=headers, allow_redirects=True, timeout=10)
        return response.url
    except Exception:
        try:
            response = requests.get(url, headers=headers, allow_redirects=True, timeout=10)
            return response.url
        except Exception:
            return url

def load_processed_invites():
    if PROCESSED_PATH.exists():
        try:
            return set(json.loads(PROCESSED_PATH.read_text()))
        except Exception:
            return set()
    return set()

def save_processed_invite(msg_key):
    processed = load_processed_invites()
    processed.add(msg_key)
    PROCESSED_PATH.write_text(json.dumps(list(processed), indent=2))

import fcntl
import threading

db_lock = threading.Lock()

def clean_name_for_search(full_name):
    # Remove special characters to avoid IMAP search encoding issues
    cleaned = full_name.replace("‘", "").replace("’", "").replace("'", "").replace("`", "")
    parts = [p.strip() for p in cleaned.split() if p.strip()]
    if len(parts) >= 2:
        return parts[0], parts[1] # Lastname, Firstname
    elif len(parts) == 1:
        return parts[0], ""
    return "", ""

def search_invite_for_email(mail, email_addr, student_name=""):
    folders = ["INBOX", "[Gmail]/Spam", "[Gmail]/All Mail"]
    for folder in folders:
        try:
            status, _ = mail.select(f'"{folder}"', readonly=True)
            if status != "OK":
                status, _ = mail.select(folder, readonly=True)
                if status != "OK":
                    continue
            
            # 1. Try searching by email_addr
            msg_ids = []
            status, data = mail.search(None, f'(TO "{email_addr}")')
            if status == "OK" and data[0]:
                msg_ids = data[0].split()
            
            # 2. Try searching by student name if email search fails
            if not msg_ids and student_name:
                last_name, first_name = clean_name_for_search(student_name)
                if last_name and first_name:
                    status, data = mail.search(None, f'TEXT "{last_name}" TEXT "{first_name}"')
                    if status == "OK" and data[0]:
                        msg_ids = data[0].split()
            
            if not msg_ids:
                continue
                
            # Fetch and check the latest message first
            for msg_id in reversed(msg_ids):
                status, msg_data = mail.fetch(msg_id, "(RFC822)")
                if status != "OK":
                    continue

                raw_email = msg_data[0][1]
                msg = email.message_from_bytes(raw_email)

                subject = msg.get("Subject") or ""
                # We only want invitation emails
                if "invited" not in subject.lower() and "invites you" not in subject.lower() and "invitation" not in subject.lower():
                    continue

                body = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        content_type = part.get_content_type()
                        content_disp = str(part.get("Content-Disposition"))
                        if content_type == "text/html" and "attachment" not in content_disp:
                            body = part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", errors="ignore")
                            break
                        elif content_type == "text/plain" and "attachment" not in content_disp:
                            body = part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", errors="ignore")
                            break
                else:
                    body = msg.get_payload(decode=True).decode(msg.get_content_charset() or "utf-8", errors="ignore")

                all_matches = re.findall(r'https?://[^\s\"\'\>]+', body)
                invite_url = None

                for raw_url in all_matches:
                    clean_url = "".join(raw_url.split()).replace("&amp;", "&").rstrip(".,;!?'\"<>#")
                    if "coursera.org" in clean_url and "invitationToken=" in clean_url:
                        invite_url = clean_url
                        break

                if not invite_url:
                    for raw_url in all_matches:
                        clean_url = "".join(raw_url.split()).replace("&amp;", "&").rstrip(".,;!?'\"<>#")
                        if "link.coursera.org/f/" in clean_url:
                            try:
                                resolved = resolve_redirect(clean_url)
                                if "invitationToken=" in resolved:
                                    invite_url = resolved
                                    break
                            except Exception:
                                pass

                if invite_url:
                    return invite_url
        except Exception as e:
            print(f"      [IMAP Search Error in folder {folder} for {email_addr}]: {e}")
    return None

def update_excel_link(passport, invite_url, email_addr):
    EXCEL_PATH = Path("Names.xlsx")
    if not EXCEL_PATH.exists():
        return
    excel_lock_path = EXCEL_PATH.with_suffix(".xlsx.lock")
    try:
        with open(excel_lock_path, "w") as lock_f:
            fcntl.flock(lock_f, fcntl.LOCK_EX)
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb["Talabalar"]
            
            # Ensure headers in F and G
            ws.cell(row=1, column=6, value="Invite Link")
            ws.cell(row=1, column=7, value="Email")

            # Map passport to sheet row
            found_row = -1
            for r in range(2, ws.max_row + 1):
                pass_val = str(ws.cell(row=r, column=2).value or "").strip()
                if pass_val == passport:
                    found_row = r
                    break
                    
            if found_row != -1:
                ws.cell(row=found_row, column=6, value=invite_url)
                ws.cell(row=found_row, column=7, value=email_addr)
                wb.save(EXCEL_PATH)
                print(f"💾 Synchronized invite link and email to Names.xlsx for Passport {passport} at row {found_row}")
    except Exception as e:
        print(f"⚠️ Error updating Excel for Passport {passport}: {e}")

def write_invite_to_db_threadsafe(passport, invite_url, retries=10, delay=1):
    lock_path = DB_PATH.with_suffix(".json.lock")
    with db_lock:
        for attempt in range(retries):
            try:
                with open(lock_path, "w") as lock_f:
                    fcntl.flock(lock_f, fcntl.LOCK_EX)
                    db = []
                    if DB_PATH.exists():
                        try:
                            db = json.loads(DB_PATH.read_text())
                        except Exception:
                            pass
                    updated = False
                    student_name = ""
                    for item in db:
                        if item.get("document") == passport:
                            item["invite_url"] = invite_url
                            student_name = item.get("name") or "Student"
                            updated = True
                            break
                    if updated:
                        DB_PATH.write_text(json.dumps(db, ensure_ascii=False, indent=2))
                        print(f"✨ Saved invite link for {student_name} to database")
                        # Sync to Excel in background
                        email_addr = next((item.get("email") for item in db if item.get("document") == passport), "")
                        update_excel_link(passport, invite_url, email_addr)
                        return True
            except Exception as err:
                print(f"⚠️ DB write attempt {attempt+1} failed: {err}. Retrying in {delay}s...")
                time.sleep(delay)
        return False

def catch_up_links(mail, db):
    print("\n" + "=" * 60)
    print("Checking for missing invite links for students 1 to 5,000...")
    print("=" * 60)
    
    # Students 1 to 5,000 correspond to rows 3 to 5002
    target_entries = []
    for entry in db:
        row = entry.get("row")
        if row is not None and 3 <= row <= 5002:
            email_val = str(entry.get("email") or "").strip().lower()
            invite_url = entry.get("invite_url")
            if email_val and not invite_url:
                target_entries.append(entry)

    if not target_entries:
        print("All students in rows 3 to 5002 already have invite links or don't have registered emails.")
        print("=" * 60 + "\n")
        return

    print(f"Found {len(target_entries)} students registered but missing invite links in rows 3 to 5002.")
    print("Going back to fetch them from Gmail IMAP...")
    
    success_count = 0
    for entry in target_entries:
        email_addr = entry["email"]
        passport = entry["document"]
        student_name = entry.get("name") or "Student"
        row_idx = entry["row"]
        
        print(f"🔍 [{success_count+1}/{len(target_entries)}] Searching Gmail for {student_name} ({email_addr}) at row {row_idx}...")
        invite_url = search_invite_for_email(mail, email_addr, student_name)
        if invite_url:
            print(f"✨ Found invite link for {student_name}: {invite_url}")
            # Update DB
            write_invite_to_db_threadsafe(passport, invite_url)
            success_count += 1
            # Sleep slightly to avoid IMAP hammer
            time.sleep(0.5)
        else:
            print(f"❌ No invite email found for {student_name} ({email_addr})")
            
    print(f"Catch-up complete! Successfully retrieved {success_count} missing invite links.")
    print("=" * 60 + "\n")

def process_emails(mail, email_to_passport, results, processed_ids):
    folders = ["INBOX", "[Gmail]/Spam", "[Gmail]/All Mail"]
    for folder in folders:
        try:
            status, _ = mail.select(f'"{folder}"', readonly=True)
            if status != "OK":
                status, _ = mail.select(folder, readonly=True)
                if status != "OK":
                    continue
            
            status, data = mail.search(None, '(FROM "coursera.org")')
            if status != "OK" or not data[0]:
                continue
                
            msg_ids = data[0].split()
            # Scan last 300 messages to catch older links
            for msg_id in reversed(msg_ids[-300:]):
                status, msg_data = mail.fetch(msg_id, "(RFC822)")
                if status != "OK":
                    continue

                raw_email = msg_data[0][1]
                msg = email.message_from_bytes(raw_email)

                msg_id_header = msg.get("Message-ID") or ""
                msg_key = msg_id_header.strip() if msg_id_header else f"{folder}:{msg_id.decode('utf-8')}"
                if msg_key in processed_ids:
                    continue

                subject = msg.get("Subject") or ""
                if "invited" not in subject.lower() and "invites you" not in subject.lower() and "invitation" not in subject.lower():
                    processed_ids.add(msg_key)
                    save_processed_invite(msg_key)
                    continue

                to_header = msg.get("To") or ""
                emails_found = re.findall(r'[\w\.-]+@[\w\.-]+', to_header)
                to_email = emails_found[0].strip().lower() if emails_found else ""

                if not to_email:
                    processed_ids.add(msg_key)
                    save_processed_invite(msg_key)
                    continue

                passport = email_to_passport.get(to_email)
                if not passport:
                    continue
                
                body = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        content_type = part.get_content_type()
                        content_disp = str(part.get("Content-Disposition"))
                        if content_type == "text/html" and "attachment" not in content_disp:
                            body = part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", errors="ignore")
                            break
                        elif content_type == "text/plain" and "attachment" not in content_disp:
                            body = part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", errors="ignore")
                else:
                    body = msg.get_payload(decode=True).decode(msg.get_content_charset() or "utf-8", errors="ignore")

                all_matches = re.findall(r'https?://[^\s\"\'\>]+', body)
                invite_url = None

                for raw_url in all_matches:
                    clean_url = "".join(raw_url.split()).replace("&amp;", "&").rstrip(".,;!?'\"<>#")
                    if "coursera.org" in clean_url and "invitationToken=" in clean_url:
                        invite_url = clean_url
                        break

                if not invite_url:
                    for raw_url in all_matches:
                        clean_url = "".join(raw_url.split()).replace("&amp;", "&").rstrip(".,;!?'\"<>#")
                        if "link.coursera.org/f/" in clean_url:
                            try:
                                resolved = resolve_redirect(clean_url)
                                if "invitationToken=" in resolved:
                                    invite_url = resolved
                                    break
                            except Exception:
                                pass

                if invite_url:
                    success = write_invite_to_db_threadsafe(passport, invite_url)
                    if success:
                        processed_ids.add(msg_key)
                        save_processed_invite(msg_key)
                else:
                    processed_ids.add(msg_key)
                    save_processed_invite(msg_key)
        except Exception as e:
            print(f"Error processing emails in folder {folder}: {e}")

def main():
    print("=" * 60)
    print("Coursera Invitation Link Extractor Daemon Started")
    print("Waiting for new invite emails...")
    print("=" * 60)

    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(GMAIL_USER, GMAIL_APP_PASS)

    # Initial catch-up on startup
    try:
        if DB_PATH.exists():
            lock_path = DB_PATH.with_suffix(".json.lock")
            with open(lock_path, "w") as lock_f:
                fcntl.flock(lock_f, fcntl.LOCK_EX)
                db = json.loads(DB_PATH.read_text())
            catch_up_links(mail, db)
    except Exception as catch_err:
        print(f"⚠️ Error during initial catch-up: {catch_err}")

    try:
        while True:
            try:
                if not DB_PATH.exists():
                    time.sleep(10)
                    continue

                processed_ids = load_processed_invites()
                
                # Read DB with process lock
                lock_path = DB_PATH.with_suffix(".json.lock")
                with open(lock_path, "w") as lock_f:
                    fcntl.flock(lock_f, fcntl.LOCK_EX)
                    db = json.loads(DB_PATH.read_text())

                email_to_passport = {}
                for entry in db:
                    email_val = str(entry.get("email") or "").strip().lower()
                    doc_val = str(entry.get("document") or "").strip()
                    if email_val and doc_val:
                        email_to_passport[email_val] = doc_val

                process_emails(mail, email_to_passport, db, processed_ids)

            except Exception as loop_err:
                print(f"Error in polling loop: {loop_err}")
                import traceback
                traceback.print_exc()
                try:
                    mail.logout()
                except Exception:
                    pass
                mail = imaplib.IMAP4_SSL("imap.gmail.com")
                mail.login(GMAIL_USER, GMAIL_APP_PASS)

            time.sleep(30)

    except KeyboardInterrupt:
        print("\nStopping...")
    finally:
        try:
            mail.logout()
        except Exception:
            pass

if __name__ == "__main__":
    main()
