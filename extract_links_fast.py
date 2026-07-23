# extract_links_fast.py
import os
import re
import json
import imaplib
import email
from email.header import decode_header
import requests
import openpyxl
import fcntl
import time
import socket
import concurrent.futures
from pathlib import Path

socket.setdefaulttimeout(15)

EXCEL_PATH = Path("Names.xlsx")
DB_PATH = Path("Names_db.json")
GMAIL_USER = "qwertyuioplkjhgfdsazxcvbnmhrh@gmail.com"
GMAIL_APP_PASS = "feykxtyuitmnjevn"
excel_lock_path = Path("excel.lock")

def get_decoded_header(header_value):
    if not header_value:
        return ""
    try:
        decoded_parts = decode_header(header_value)
        header_text = ""
        for part, encoding in decoded_parts:
            if isinstance(part, bytes):
                header_text += part.decode(encoding or "utf-8", errors="ignore")
            else:
                header_text += part
        return header_text
    except Exception:
        return str(header_value)

def normalize_email_for_match(email_addr):
    if not email_addr or "@" not in email_addr:
        return ""
    return email_addr.strip().lower()

def resolve_redirect(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    try:
        response = requests.head(url, headers=headers, allow_redirects=True, timeout=10)
        return response.url
    except Exception:
        try:
            response = requests.get(url, headers=headers, allow_redirects=True, timeout=10)
            return response.url
        except Exception:
            return url

def main():
    print("📖 Loading target list from Names.xlsx and Names_db.json...")
    
    # Load JSON DB
    db_data = []
    if DB_PATH.exists():
        try:
            db_data = json.loads(DB_PATH.read_text())
            print(f"Loaded {len(db_data)} students from Names_db.json.")
        except Exception as e:
            print(f"⚠️ Could not load database JSON: {e}")

    # Load Excel sheet
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Talabalar"]

    email_to_student = {}
    
    # We load students from row 3 to 5200
    for r_idx in range(3, 5200):
        passport = str(ws.cell(row=r_idx, column=2).value or "").strip()
        if not passport:
            continue

        existing_link = ws.cell(row=r_idx, column=6).value
        has_link = existing_link and "coursera.org" in str(existing_link)

        email_addr = str(ws.cell(row=r_idx, column=7).value or "").strip().lower()
        
        # Target students: registered email but no invite link yet in Excel
        if email_addr and not has_link:
            norm_email = normalize_email_for_match(email_addr)
            student_record = {
                "row": r_idx,
                "passport": passport,
                "name": str(ws.cell(row=r_idx, column=1).value or "").strip(),
                "email": email_addr,
                "norm_email": norm_email
            }
            email_to_student[norm_email] = student_record

    print(f"Loaded {len(email_to_student)} target students missing invite links in Excel.")
    if not email_to_student:
        print("All target students already have invite links in Excel. Exiting.")
        return

    # Connect to Gmail IMAP
    print("Connecting to Gmail IMAP...", flush=True)
    mail = imaplib.IMAP4_SSL("imap.gmail.com", timeout=15)
    mail.login(GMAIL_USER, GMAIL_APP_PASS)
    
    # We will search the INBOX first, as all Coursera invites are in INBOX
    folder = "INBOX"
    print(f"\n📂 Scanning folder: {folder}...", flush=True)
    status, _ = mail.select(f'"{folder}"', readonly=True)
    if status != "OK":
        status, _ = mail.select(folder, readonly=True)
        if status != "OK":
            print(f"❌ Failed to select {folder}. Exiting.")
            mail.logout()
            return

    # Search for all messages since July 15, 2026
    status, data = mail.search(None, '(SINCE "15-Jul-2026")')
    if status != "OK" or not data[0]:
        print(f"No messages found in {folder} since 15-Jul-2026.")
        mail.logout()
        return

    msg_ids = data[0].split()
    print(f"Found {len(msg_ids)} messages in {folder}. Fetching To/Subject headers in bulk...", flush=True)

    # Fetch To and Subject in batches of 1000 message IDs to stay well within IMAP command limits
    header_batch_size = 1000
    msg_id_to_student = {}
    
    for i in range(0, len(msg_ids), header_batch_size):
        batch = msg_ids[i:i+header_batch_size]
        batch_str = b",".join(batch)
        print(f"  ✉️ Fetching headers batch {i//header_batch_size + 1}/{(len(msg_ids)+header_batch_size-1)//header_batch_size}...", flush=True)
        status, fetch_data = mail.fetch(batch_str, "(BODY[HEADER.FIELDS (TO SUBJECT)])")
        if status != "OK":
            continue
            
        for item in fetch_data:
            if isinstance(item, tuple):
                try:
                    m_id = item[0].split()[0].decode()
                    msg = email.message_from_bytes(item[1])
                    
                    subject = get_decoded_header(msg.get("Subject"))
                    if "invited" not in subject.lower() and "invitation" not in subject.lower():
                        continue

                    to_header = get_decoded_header(msg.get("To"))
                    to_emails = re.findall(r'[\w\.-]+@[\w\.-]+', to_header)
                    if not to_emails:
                        continue
                    to_email = to_emails[0].strip().lower()
                    norm_to = normalize_email_for_match(to_email)
                    
                    if norm_to in email_to_student:
                        msg_id_to_student[m_id] = email_to_student[norm_to]
                except Exception:
                    pass

    print(f"Found {len(msg_id_to_student)} emails matching our target students.", flush=True)
    if not msg_id_to_student:
        print("No matching invite emails found. Exiting.")
        mail.logout()
        return

    # Fetch bodies of matched messages in batches of 10
    matched_ids = list(msg_id_to_student.keys())
    batch_size = 10
    updates_found = {}

    print(f"Fetching bodies for {len(matched_ids)} matching messages...", flush=True)
    for i in range(0, len(matched_ids), batch_size):
        batch = matched_ids[i:i+batch_size]
        batch_ids = b",".join(m.encode() for m in batch)
        batch_idx = i // batch_size + 1
        total_batches = (len(matched_ids) + batch_size - 1) // batch_size
        
        # Retry loop for quota limits and connection drops
        body_data = None
        for attempt in range(1, 4):
            print(f"  ✉️ Fetching bodies batch {batch_idx}/{total_batches} (Attempt {attempt}/3)...", flush=True)
            try:
                status, body_data = mail.fetch(batch_ids, "(BODY.PEEK[TEXT])")
                if status == "OK" and body_data:
                    break
                else:
                    print(f"  ⚠️ Fetch failed with status {status}. Retrying in 5s...", flush=True)
            except Exception as e:
                err_str = str(e)
                print(f"  ⚠️ Fetch error: {err_str}.", flush=True)
                
                # Check if connection is closed or dead, and reconnect
                is_dead = "EOF" in err_str or "closed" in err_str or "broken" in err_str or "connection" in err_str.lower() or "timeout" in err_str.lower() or "timed out" in err_str.lower()
                if is_dead:
                    print("  🔄 Connection appears dead. Reconnecting and logging in...", flush=True)
                    try:
                        try:
                            mail.logout()
                        except Exception:
                            pass
                        mail = imaplib.IMAP4_SSL("imap.gmail.com", timeout=15)
                        mail.login(GMAIL_USER, GMAIL_APP_PASS)
                        mail.select(f'"{folder}"', readonly=True)
                        print("  ✅ Reconnected successfully.", flush=True)
                    except Exception as reconnect_err:
                        print(f"  ❌ Reconnection failed: {reconnect_err}", flush=True)
                
                if "OVERQUOTA" in err_str or "exceeded" in err_str or is_dead:
                    print("  Sleeping 10s to cool down...", flush=True)
                    time.sleep(10)
            time.sleep(5)
        
        if not body_data:
            print(f"  ❌ Failed batch {batch_idx} after 3 attempts. Bypassing.", flush=True)
            continue

        # Parse bodies
        for part in body_data:
            if isinstance(part, tuple):
                try:
                    m_id = part[0].split()[0].decode()
                    student = msg_id_to_student[m_id]
                    
                    import quopri
                    decoded_bytes = quopri.decodestring(part[1])
                    body_text = decoded_bytes.decode("utf-8", errors="ignore")
                    
                    # Extract Coursera link
                    invite_link = ""
                    button_pattern = r'(?:Join now|Join Program|Start learning|Accept invitation|Accept Invitation)\s*<(https://[^\s"<>]+coursera\.org/[^>]+)>'
                    match = re.search(button_pattern, body_text, re.IGNORECASE)
                    if match:
                        l = match.group(1)
                        parts_url = l.split("?")
                        if len(parts_url) > 1:
                            invite_link = parts_url[0] + "?" + parts_url[1].split()[0]
                        else:
                            invite_link = l
                        invite_link = invite_link.rstrip('>').rstrip('"').rstrip(')').rstrip('.')
                    
                    if invite_link:
                        updates_found[student["row"]] = {
                            "invite_link": invite_link,
                            "email": student["email"],
                            "passport": student["passport"],
                            "name": student["name"]
                        }
                except Exception as e:
                    pass
        
        # Be polite to Gmail IMAP server rate limits
        time.sleep(0.5)

    mail.logout()
    print(f"Successfully parsed {len(updates_found)} invite links from email bodies. Resolving redirects in parallel...", flush=True)

    # Resolve redirects in parallel
    resolved_updates = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
        future_to_row = {
            executor.submit(resolve_redirect, item["invite_link"]): row
            for row, item in updates_found.items()
        }
        for future in concurrent.futures.as_completed(future_to_row):
            row = future_to_row[future]
            try:
                resolved_url = future.result()
                orig_item = updates_found[row]
                resolved_updates[row] = {
                    "invite_url": resolved_url,
                    "email": orig_item["email"],
                    "passport": orig_item["passport"],
                    "name": orig_item["name"]
                }
                print(f"  ✨ Resolved Row {row}: {orig_item['name']} -> {resolved_url}", flush=True)
            except Exception as e:
                print(f"  ⚠️ Error resolving redirect for row {row}: {e}", flush=True)

    print(f"\n💾 Saving {len(resolved_updates)} resolved links to Names.xlsx and Names_db.json...", flush=True)

    # Write updates to Excel
    if resolved_updates:
        try:
            with open(excel_lock_path, "w") as lock_f:
                fcntl.flock(lock_f, fcntl.LOCK_EX)
                wb_w = openpyxl.load_workbook(EXCEL_PATH)
                ws_w = wb_w["Talabalar"]
                for row_idx, data in resolved_updates.items():
                    ws_w.cell(row=row_idx, column=6, value=data["invite_url"])
                    ws_w.cell(row=row_idx, column=7, value=data["email"])
                wb_w.save(EXCEL_PATH)
                fcntl.flock(lock_f, fcntl.LOCK_UN)
            print(f"  ✅ Names.xlsx updated with {len(resolved_updates)} new links.", flush=True)
        except Exception as err:
            print(f"  ❌ Failed to write Names.xlsx: {err}", flush=True)

        # Write updates to JSON DB
        if DB_PATH.exists():
            lock_path = DB_PATH.with_suffix(".json.lock")
            try:
                with open(lock_path, "w") as lock_f2:
                    fcntl.flock(lock_f2, fcntl.LOCK_EX)
                    db = json.loads(DB_PATH.read_text())
                    passport_updates = {data["passport"]: data["invite_url"] for data in resolved_updates.values()}
                    for item in db:
                        doc = item.get("document")
                        if doc in passport_updates:
                            item["invite_url"] = passport_updates[doc]
                            item["activated"] = True
                    DB_PATH.write_text(json.dumps(db, ensure_ascii=False, indent=2))
                    fcntl.flock(lock_f2, fcntl.LOCK_UN)
                print(f"  ✅ Names_db.json updated.", flush=True)
            except Exception as err:
                print(f"  ❌ Failed to write Names_db.json: {err}", flush=True)

    print(f"\nDone! Successfully resolved and synchronized {len(resolved_updates)} invitation links.", flush=True)

if __name__ == "__main__":
    main()
