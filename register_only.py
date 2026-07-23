# register_only.py
import json
import random
import re
import time
import requests
import threading
import concurrent.futures
import fcntl
import openpyxl
from pathlib import Path

# Config
DB_PATH = Path("Names_db.json")
USED_EMAILS_PATH = Path("used_emails.json")
BASE_URL = "https://aileaders.uz"
GMAIL_USER = "qwertyuioplkjhgfdsazxcvbnmhrh@gmail.com"
CONCURRENCY = 25
LIMIT = 5102

HEADERS = {
    "accept": "*/*",
    "accept-language": "en-US,en;q=0.9,uz;q=0.8",
    "content-type": "application/json",
    "origin": "https://aileaders.uz",
    "priority": "u=1, i",
    "referer": "https://aileaders.uz/auth/register",
    "sec-ch-ua": "\"Not;A=Brand\";v=\"8\", \"Chromium\";v=\"150\", \"Google Chrome\";v=\"150\"",
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": "\"Linux\"",
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-origin",
    "user-agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36",
    "Cookie": "HWWAFSESTIME=1784291135540; HWWAFSESID=45eff767bfb2890ea5",
}

db_lock = threading.Lock()

def load_used_emails():
    if USED_EMAILS_PATH.exists():
        try:
            return set(json.loads(USED_EMAILS_PATH.read_text()))
        except Exception:
            return set()
    return set()

def generate_dot_alias(base_username, used_emails):
    L = len(base_username)
    while True:
        parts = [base_username[0]]
        for i in range(1, L):
            if random.choice([True, False]):
                parts.append(".")
            parts.append(base_username[i])
        email_addr = "".join(parts) + "@gmail.com"
        if email_addr not in used_emails:
            return email_addr

def generate_dot_alias_threadsafe(base_username):
    with db_lock:
        used = []
        if USED_EMAILS_PATH.exists():
            try:
                used = json.loads(USED_EMAILS_PATH.read_text())
            except Exception:
                pass
        used_emails = set(used)
        email_addr = generate_dot_alias(base_username, used_emails)
        used_emails.add(email_addr)
        USED_EMAILS_PATH.write_text(json.dumps(list(used_emails), indent=2))
        return email_addr

def first_api_call(session, document, dob):
    url = f"{BASE_URL}/api/public/info/individual"
    params = {"document": document, "dob": dob, "occupation": "student"}
    response = session.post(url, params=params, headers=HEADERS, data="")
    response.raise_for_status()
    return response.json() if response.content else {}

def second_api_call(session, email_addr, document, dob, phone):
    valid_phone = "+998995337221"
    url = f"{BASE_URL}/api/registration/form"
    payload = {
        "email": email_addr,
        "employment_type": "student",
        "metrika": None,
        "passport": {"document": document, "dob": dob},
        "password": document,
        "phone": valid_phone,
    }
    response = session.post(url, headers=HEADERS, json=payload)
    response.raise_for_status()
    return response.json()

def delete_account(session, document, dob):
    url = f"{BASE_URL}/api/profile/delete-account"
    headers = HEADERS.copy()
    headers["content-type"] = "application/x-www-form-urlencoded"
    headers["referer"] = f"{BASE_URL}/auth/delete_account"
    payload = f"document={document}&dob={dob}"
    response = session.delete(url, headers=headers, data=payload)
    response.raise_for_status()
    return response.json() if response.content else {}

def sync_single_to_excel(passport, email_addr):
    EXCEL_PATH = Path("Names.xlsx")
    if not EXCEL_PATH.exists():
        return
    excel_lock_path = EXCEL_PATH.with_suffix(".xlsx.lock")
    try:
        with open(excel_lock_path, "w") as lock_f:
            fcntl.flock(lock_f, fcntl.LOCK_EX)
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb["Talabalar"]
            
            # Map passport to sheet row
            found_row = -1
            for r in range(2, ws.max_row + 1):
                pass_val = str(ws.cell(row=r, column=2).value or "").strip()
                if pass_val == passport:
                    found_row = r
                    break
                    
            if found_row != -1:
                ws.cell(row=found_row, column=7, value=email_addr)
                wb.save(EXCEL_PATH)
                print(f"      💾 Excel: Row {found_row} Column G updated with {email_addr}")
            fcntl.flock(lock_f, fcntl.LOCK_UN)
    except Exception as e:
        print(f"⚠️ Excel sync warning for Passport {passport}: {e}")

def save_result_threadsafe(entry):
    lock_path = DB_PATH.with_suffix(".json.lock")
    with db_lock:
        try:
            with open(lock_path, "w") as lock_f:
                fcntl.flock(lock_f, fcntl.LOCK_EX)
                db = []
                if DB_PATH.exists():
                    try:
                        db = json.loads(DB_PATH.read_text())
                    except Exception:
                        pass
                for item in db:
                    if item.get("document") == entry.get("document"):
                        item.update(entry)
                        break
                DB_PATH.write_text(json.dumps(db, ensure_ascii=False, indent=2))
                fcntl.flock(lock_f, fcntl.LOCK_UN)
                print(f"      💾 Saved registration details to Names_db.json.")
                sync_single_to_excel(entry["document"], entry["email"])
        except Exception as e:
            print(f"Error saving to DB: {e}")

def register_worker(worker_id, row, base_username):
    orig_row = row["original_row_index"]
    print(f"🔹 [Worker {worker_id}] Starting Excel Row {orig_row}: Doc={row['document']}")
    
    session = requests.Session()
    try:
        session.get(BASE_URL, headers={"User-Agent": HEADERS["user-agent"]})
    except Exception as e:
        pass

    success = False
    attempts = 0
    max_attempts = 3
    last_email_used = None

    try:
        while attempts < max_attempts and not success:
            attempts += 1
            email_addr = generate_dot_alias_threadsafe(base_username)
            last_email_used = email_addr
            print(f"🔹 [Worker {worker_id}] Row {orig_row} Attempt {attempts}/{max_attempts}: alias={email_addr}")

            try:
                first_api_call(session, row["document"], row["dob"])
                time.sleep(0.1)

                resp = second_api_call(session, email_addr, row["document"], row["dob"], row["phone"])
                res_code = resp.get("result", {}).get("code")

                if res_code != "ok":
                    if res_code == "passport_is_already_in_use":
                        print(f"🔹 [Worker {worker_id}] Row {orig_row} passport already in use. Calling delete_account...")
                        try:
                            delete_account(session, row["document"], row["dob"])
                        except Exception as del_e:
                            pass
                        time.sleep(0.1)
                        continue
                    elif "email" in str(res_code):
                        continue
                    else:
                        print(f"❌ [Worker {worker_id}] Row {orig_row} registration rejected: {res_code}")
                        break

                # Get JWT login token
                login_resp = session.post(f"{BASE_URL}/api/authorization/login", headers={
                    **HEADERS,
                    "referer": f"{BASE_URL}/auth/login",
                }, json={
                    "login": email_addr,
                    "password": row["document"]
                })
                login_resp.raise_for_status()
                login_data = login_resp.json()
                token = login_data.get("content", {}).get("token")
                
                if not token:
                    print(f"🔹 [Worker {worker_id}] Row {orig_row} failed to get login token.")
                    continue

                # Trigger email verify link
                verify_resp = session.post(f"{BASE_URL}/api/profile/verify-email?email={email_addr}", headers={
                    **HEADERS,
                    "Authorization": f"Bearer {token}"
                })
                verify_resp.raise_for_status()

                # Registration successful, email triggered! Save it!
                result_entry = {
                    "document": row["document"],
                    "email": email_addr,
                    "activated": False,      # Will be activated later in Phase 2
                    "invite_url": None
                }
                save_result_threadsafe(result_entry)
                success = True
                print(f"✅ [Worker {worker_id}] Row {orig_row} REGISTERED & EMAIL SENT!")

            except Exception as e:
                print(f"🔹 [Worker {worker_id}] Row {orig_row} error: {e}")
                time.sleep(0.2)

        if not success:
            print(f"❌ [Worker {worker_id}] Row {orig_row} failed after {attempts} attempts.")

    except Exception as e:
        print(f"Worker exception: {e}")

def main():
    print("=" * 60)
    print("AIleaders Fast Register-Only Script Started")
    print(f"Running with concurrency = {CONCURRENCY}")
    print("=" * 60)

    # Read database rows
    if not DB_PATH.exists():
        print("Database not found!")
        return

    all_rows = json.loads(Path(DB_PATH).read_text())
    
    pending_rows = []
    for r in all_rows:
        r["original_row_index"] = r["row"]
        if r.get("row") > LIMIT:
            continue
        # Only process rows that do not have email registered yet
        if not r.get("email"):
            pending_rows.append(r)

    print(f"Total rows up to {LIMIT}: {len(all_rows)}")
    print(f"Pending registrations: {len(pending_rows)}")

    if not pending_rows:
        print("All rows already registered! Nothing to do.")
        return

    base_username = GMAIL_USER.split("@")[0]

    with concurrent.futures.ThreadPoolExecutor(max_workers=CONCURRENCY) as executor:
        futures = {}
        for worker_id, row in enumerate(pending_rows, start=1):
            w_id = ((worker_id - 1) % CONCURRENCY) + 1
            f = executor.submit(register_worker, w_id, row, base_username)
            futures[f] = row

        for future in concurrent.futures.as_completed(futures):
            row = futures[future]
            try:
                future.result()
            except Exception as e:
                print(f"❌ Exception in Row {row['original_row_index']}: {e}")

    print("\n✅ Fast registration phase complete!")

if __name__ == "__main__":
    main()
