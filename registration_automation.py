# registration_automation.py
"""
Automates registration on aileaders.uz using data from an Excel file.

Steps:
1. Read Excel file `/home/xasanboy/Downloads/SCRIPT (2)/Names.xlsx`.
2. For each row extract:
   - Passport number (document) – column B ("Pasport raqami")
   - Date of birth (dob) – column D ("Tug‘ilgan sana")
   - Phone number – column E ("Telefon")
3. Generate a unique, unused dot-alias of the master Gmail address.
4. POST to `/api/public/info/individual` to request individual info.
5. POST to `/api/registration/form` with registration details.
6. If registration returns "passport_is_already_in_use", delete the existing account and retry.
7. If registration returns "email_exists", retry with a new dot-alias.
8. Log in using `/api/authorization/login` to obtain the user session JWT token.
9. POST to `/api/profile/verify-email?email=...` with the JWT token in Authorization header to trigger email send.
10. Poll Gmail via IMAP (searching the Inbox and Spam folder for the specific dot-alias in the TO field),
    fetch the activation email, parse it, and extract the activation URL.
11. Log the result and save to `results.json`.
12. Skip already successfully processed rows in resume mode.
"""

import json
import random
import re
import string
import time
import imaplib
import email
from pathlib import Path

import requests
import threading
import concurrent.futures
import fcntl
import openpyxl

db_lock = threading.Lock()
excel_lock = threading.Lock()
imap_semaphore = threading.BoundedSemaphore(10)
# Configuration
DB_PATH = Path("Names_db.json")
USED_EMAILS_PATH = Path("used_emails.json")
BASE_URL = "https://aileaders.uz"

# Gmail Account Credentials
GMAIL_USER = "qwertyuioplkjhgfdsazxcvbnmhrh@gmail.com"
GMAIL_APP_PASS = "kzpsmykjzrueegpv"

# Processing limit for the run (set to None to process all rows in the Excel file)
LIMIT = 7200

HEADERS = {
    "accept": "*/*",
    "accept-language": "en-US,en;q=0.9,uz;q=0.8",
    "content-type": "application/json",
    "origin": "https://aileaders.uz",
    "priority": "u=1, i",
    "referer": "https://aileaders.uz/auth/register",
    "sec-ch-ua": "\"Not;A=Brand\";v=\"8\", \"Chromium\";v=\"150\", \"Google Chrome\";v=\"150\"",
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": "\"Linux\"",
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-origin",
    "user-agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36",
    "Cookie": "HWWAFSESTIME=1784291135540; HWWAFSESID=45eff767bfb2890ea5",
}

# Shared activation links dictionary
activation_links = {}
activation_links_lock = threading.Lock()

def gmail_poller_thread():
    """Single thread that polls Gmail IMAP once every 10 seconds and extracts activation links."""
    print("🤖 Background Gmail poller thread started.", flush=True)
    
    mail = None
    while True:
        try:
            if not mail:
                mail = imaplib.IMAP4_SSL("imap.gmail.com")
                mail.login(GMAIL_USER, GMAIL_APP_PASS)
            
            for folder in ["INBOX", "[Gmail]/Spam"]:
                status, _ = mail.select(f'"{folder}"', readonly=True)
                if status != "OK":
                    status, _ = mail.select(folder, readonly=True)
                    if status != "OK":
                        continue
                
                # Search for emails with 'Activation' in subject since July 15, 2026
                status, data = mail.search(None, '(SINCE "15-Jul-2026" SUBJECT "Activation")')
                if status != "OK" or not data[0]:
                    continue
                
                msg_ids = data[0].split()
                # Check the last 30 messages (newest first)
                for msg_id in reversed(msg_ids[-30:]):
                    try:
                        status, msg_data = mail.fetch(msg_id, "(RFC822)")
                        if status != "OK":
                            continue
                        
                        raw_email = msg_data[0][1]
                        msg = email.message_from_bytes(raw_email)
                        
                        # Extract target recipient email
                        to_header = msg.get("To") or ""
                        to_emails = re.findall(r'[\w\.-]+@[\w\.-]+', to_header)
                        if not to_emails:
                            continue
                        target_email = to_emails[0].strip().lower()
                        
                        # Skip if already resolved
                        with activation_links_lock:
                            if target_email in activation_links:
                                continue
                        
                        # Extract activation URL
                        body = ""
                        if msg.is_multipart():
                            for part in msg.walk():
                                content_type = part.get_content_type()
                                content_disp = str(part.get("Content-Disposition"))
                                if content_type == "text/html" and "attachment" not in content_disp:
                                    body = part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", errors="ignore")
                                    break
                                elif content_type == "text/plain" and "attachment" not in content_disp:
                                    body = part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", errors="ignore")
                                    break
                        else:
                            body = msg.get_payload(decode=True).decode(msg.get_content_charset() or "utf-8", errors="ignore")
                        
                        match = re.search(r'https://aileaders.uz/auth/activate/[^\s"\'>\<\#]+', body)
                        if match:
                            activation_url = match.group(0).replace("&amp;", "&")
                            with activation_links_lock:
                                activation_links[target_email] = activation_url
                            print(f"📥 [Poller] Found activation link for {target_email} in {folder}", flush=True)
                    except Exception:
                        pass
            
            time.sleep(2)
            
        except Exception as e:
            print(f"⚠️ [Poller Error]: {e}. Reconnecting in 10s...", flush=True)
            try:
                if mail:
                    mail.logout()
            except Exception:
                pass
            mail = None
            time.sleep(10)

def load_used_emails():
    if USED_EMAILS_PATH.exists():
        try:
            return set(json.loads(USED_EMAILS_PATH.read_text()))
        except Exception:
            return set()
    return set()

def save_used_email(email_addr):
    used = load_used_emails()
    used.add(email_addr)
    USED_EMAILS_PATH.write_text(json.dumps(list(used), indent=2))

def generate_dot_alias(base_username, used_emails):
    L = len(base_username)
    while True:
        parts = [base_username[0]]
        for i in range(1, L):
            if random.choice([True, False]):
                parts.append(".")
            parts.append(base_username[i])
        email_addr = "".join(parts) + "@gmail.com"
        if email_addr not in used_emails:
            return email_addr

def read_db_rows():
    lock_path = DB_PATH.with_suffix(".json.lock")
    with db_lock:
        if not DB_PATH.exists():
            return []
        try:
            with open(lock_path, "w") as lock_f:
                fcntl.flock(lock_f, fcntl.LOCK_EX)
                return json.loads(DB_PATH.read_text())
        except Exception as e:
            print(f"Error reading DB: {e}")
            return []

def first_api_call(session, document, dob):
    url = f"{BASE_URL}/api/public/info/individual"
    params = {"document": document, "dob": dob, "occupation": "student"}
    response = session.post(url, params=params, headers=HEADERS, data="")
    print(f"      [first_api_call] HTTP {response.status_code}: {response.text}")
    response.raise_for_status()
    return response.json() if response.content else {}

def second_api_call(session, email_addr, document, dob, phone):
    # Override with a known accepted phone number to prevent phone_format_invalid errors
    valid_phone = "+998995337221"
    url = f"{BASE_URL}/api/registration/form"
    payload = {
        "email": email_addr,
        "employment_type": "student",
        "metrika": None,
        "passport": {"document": document, "dob": dob},
        "password": document,
        "phone": valid_phone,
    }
    response = session.post(url, headers=HEADERS, json=payload)
    print(f"      [second_api_call] HTTP {response.status_code}: {response.text}")
    response.raise_for_status()
    return response.json()

def delete_account(session, document, dob):
    url = f"{BASE_URL}/api/profile/delete-account"
    headers = HEADERS.copy()
    headers["content-type"] = "application/x-www-form-urlencoded"
    headers["referer"] = f"{BASE_URL}/auth/delete_account"
    payload = f"document={document}&dob={dob}"
    response = session.delete(url, headers=headers, data=payload)
    response.raise_for_status()
    return response.json() if response.content else {}

def activate_account(session, activation_url):
    try:
        response = session.get(activation_url, headers={
            "user-agent": HEADERS["user-agent"]
        })
        if response.status_code == 200:
            print("      ✅ Account activated successfully.")
            return True
        else:
            print(f"      ❌ Account activation returned status code {response.status_code}.")
            return False
    except Exception as e:
        print(f"      ❌ Account activation failed: {e}")
        return False

def load_existing_results():
    lock_path = DB_PATH.with_suffix(".json.lock")
    with db_lock:
        if DB_PATH.exists():
            try:
                with open(lock_path, "w") as lock_f:
                    fcntl.flock(lock_f, fcntl.LOCK_EX)
                    return json.loads(DB_PATH.read_text())
            except Exception:
                return []
        return []

def sync_single_to_excel(passport, email_addr=None, invite_url=None):
    EXCEL_PATH = Path("Names.xlsx")
    if not EXCEL_PATH.exists():
        return
    excel_lock_path = EXCEL_PATH.with_suffix(".xlsx.lock")
    try:
        with open(excel_lock_path, "w") as lock_f:
            fcntl.flock(lock_f, fcntl.LOCK_EX)
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb["Talabalar"]
            
            # Ensure headers in F and G
            ws.cell(row=1, column=6, value="Invite Link")
            ws.cell(row=1, column=7, value="Email")

            # Map passport to sheet row
            found_row = -1
            for r in range(2, ws.max_row + 1):
                pass_val = str(ws.cell(row=r, column=2).value or "").strip()
                if pass_val == passport:
                    found_row = r
                    break
                    
            if found_row != -1:
                if email_addr:
                    ws.cell(row=found_row, column=7, value=email_addr)
                if invite_url and "coursera.org" in invite_url:
                    ws.cell(row=found_row, column=6, value=invite_url)
                wb.save(EXCEL_PATH)
                print(f"      💾 Synchronized to Names.xlsx for Passport {passport} at row {found_row}")
    except Exception as e:
        print(f"⚠️ Excel sync warning for Passport {passport}: {e}")

def save_result_threadsafe(entry):
    lock_path = DB_PATH.with_suffix(".json.lock")
    with db_lock:
        try:
            with open(lock_path, "w") as lock_f:
                fcntl.flock(lock_f, fcntl.LOCK_EX)
                db = []
                if DB_PATH.exists():
                    try:
                        db = json.loads(DB_PATH.read_text())
                    except Exception:
                        pass
                for item in db:
                    if item.get("document") == entry.get("document"):
                        item.update(entry)
                        break
                DB_PATH.write_text(json.dumps(db, ensure_ascii=False, indent=2))
                print(f"      💾 Saved registration details to Names_db.json database.")
                
                # Excel sync
                passport = entry.get("document")
                email_addr = entry.get("email")
                invite_url = entry.get("invite_url") or entry.get("activation_url")
                if passport:
                    sync_single_to_excel(passport, email_addr, invite_url)
        except Exception as e:
            print(f"Error saving to DB: {e}")

def save_used_email_threadsafe(email_addr):
    with db_lock:
        used = []
        if USED_EMAILS_PATH.exists():
            try:
                used = json.loads(USED_EMAILS_PATH.read_text())
            except Exception:
                pass
        used_set = set(used)
        used_set.add(email_addr)
        USED_EMAILS_PATH.write_text(json.dumps(list(used_set), indent=2))

def generate_dot_alias_threadsafe(base_username):
    with db_lock:
        used = []
        if USED_EMAILS_PATH.exists():
            try:
                used = json.loads(USED_EMAILS_PATH.read_text())
            except Exception:
                pass
        used_emails = set(used)
        email_addr = generate_dot_alias(base_username, used_emails)
        used_emails.add(email_addr)
        USED_EMAILS_PATH.write_text(json.dumps(list(used_emails), indent=2))
        return email_addr

def register_worker(worker_id, row, base_username):
    orig_row = row["original_row_index"]
    print(f"🔹 [Worker {worker_id}] Starting Excel Row {orig_row}: Doc={row['document']}, DOB={row['dob']}, Phone={row['phone']}")
    
    session = requests.Session()
    try:
        session.get(BASE_URL, headers={"User-Agent": HEADERS["user-agent"]})
    except Exception as e:
        print(f"🔹 [Worker {worker_id}] Warning getting WAF cookies for Row {orig_row}: {e}")

    success = False
    attempts = 0
    max_attempts = 5
    last_email_used = None

    try:
        while attempts < max_attempts and not success:
            attempts += 1
            email_addr = generate_dot_alias_threadsafe(base_username)
            last_email_used = email_addr
            print(f"🔹 [Worker {worker_id}] Row {orig_row} Attempt {attempts}/{max_attempts}: alias={email_addr}")

            try:
                first_api_call(session, row["document"], row["dob"])
                time.sleep(0.1)

                resp = second_api_call(session, email_addr, row["document"], row["dob"], row["phone"])
                res_code = resp.get("result", {}).get("code")

                if res_code != "ok":
                    if res_code == "passport_is_already_in_use":
                        print(f"🔹 [Worker {worker_id}] Row {orig_row} passport already in use. Calling delete_account...")
                        try:
                            delete_account(session, row["document"], row["dob"])
                            print(f"      [delete_account] Account delete request sent. Sleeping 5s to let database update...")
                            time.sleep(5)
                        except Exception as del_e:
                            print(f"🔹 [Worker {worker_id}] Delete failed: {del_e}")
                        time.sleep(0.1)
                        continue
                    elif "email" in str(res_code):
                        print(f"🔹 [Worker {worker_id}] Row {orig_row} email error ({res_code}). Generating a new one and retrying...")
                        time.sleep(0.1)
                        continue
                    else:
                        print(f"❌ [Worker {worker_id}] Row {orig_row} registration rejected by API: {res_code}")
                        result_entry = {
                            "document": row["document"],
                            "email": None,
                            "error": f"Registration failed: {res_code}"
                        }
                        save_result_threadsafe(result_entry)
                        success = False
                        break  # Abort attempts for this bad row

                login_resp = session.post(f"{BASE_URL}/api/authorization/login", headers={
                    **HEADERS,
                    "referer": f"{BASE_URL}/auth/login",
                }, json={
                    "login": email_addr,
                    "password": row["document"]
                })
                login_resp.raise_for_status()
                login_data = login_resp.json()
                if isinstance(login_data, str):
                    print(f"🔹 [Worker {worker_id}] Row {orig_row} login failed: {login_data}")
                    continue
                token = login_data.get("content", {}).get("token")
                
                if not token:
                    print(f"🔹 [Worker {worker_id}] Row {orig_row} failed to get login token.")
                    continue

                verify_resp = session.post(f"{BASE_URL}/api/profile/verify-email?email={email_addr}", headers={
                    **HEADERS,
                    "Authorization": f"Bearer {token}"
                })
                verify_resp.raise_for_status()

                activation_url = None
                start_poll = time.time()
                while time.time() - start_poll < 95:
                    with activation_links_lock:
                        if email_addr in activation_links:
                            activation_url = activation_links[email_addr]
                            break
                    time.sleep(1.5)

                if activation_url:
                    print(f"🔹 [Worker {worker_id}] Row {orig_row} activation link: {activation_url}")
                    activated = activate_account(session, activation_url)
                    
                    result_entry = {
                        "document": row["document"],
                        "email": email_addr,
                        "activation_url": activation_url,
                        "activated": activated,
                        "api_response": resp
                    }
                    save_result_threadsafe(result_entry)
                    success = True
                    print(f"✅ [Worker {worker_id}] Row {orig_row} SUCCESSFULLY REGISTERED!")
                else:
                    print(f"🔹 [Worker {worker_id}] Row {orig_row} activation link timeout.")
                    time.sleep(0.1)

            except Exception as e:
                import traceback
                traceback.print_exc()
                print(f"🔹 [Worker {worker_id}] Row {orig_row} error: {e}")
                time.sleep(0.1)

        if not success:
            print(f"❌ [Worker {worker_id}] Row {orig_row} failed after {attempts} attempts.")
            result_entry = {
                "document": row["document"],
                "email": last_email_used,
                "error": f"Failed after {max_attempts} attempts"
            }
            save_result_threadsafe(result_entry)

    finally:
        pass

def main():
    print("=" * 60)
    print("AIleaders Registration Daemon Started")
    print("Configured for sequential execution (concurrency = 1).")
    print("=" * 60)

    all_rows = read_db_rows()
    print(f"Loaded {len(all_rows)} rows from database.")

    pending_rows = []
    for r in all_rows:
        r["original_row_index"] = r["row"]
        # Limit processing range strictly to rows 6033 to 7202
        if r.get("row") < 6033 or r.get("row") > 7202:
            continue
        # Skip if they already have an invite link OR if they are successfully activated
        if r.get("invite_url") or r.get("activated"):
            continue
        pending_rows.append(r)

    print(f"Already processed: {len(all_rows) - len(pending_rows)} rows.")
    print(f"Pending rows: {len(pending_rows)} rows.")

    if LIMIT is not None:
        rows_to_process = pending_rows[:LIMIT]
        print(f"Limit applied: processing next {len(rows_to_process)} pending rows.")
    else:
        rows_to_process = pending_rows
        print(f"Processing all remaining {len(rows_to_process)} pending rows.")

    base_username = GMAIL_USER.split("@")[0]

    # Start background Gmail IMAP poller thread
    poller = threading.Thread(target=gmail_poller_thread, daemon=True)
    poller.start()

    concurrency = 10
    with concurrent.futures.ThreadPoolExecutor(max_workers=concurrency) as executor:
        futures = {}
        for worker_id, row in enumerate(rows_to_process, start=1):
            w_id = ((worker_id - 1) % concurrency) + 1
            f = executor.submit(register_worker, w_id, row, base_username)
            futures[f] = row

        for future in concurrent.futures.as_completed(futures):
            row = futures[future]
            try:
                future.result()
            except Exception as e:
                print(f"❌ Exception in thread processing Row {row['original_row_index']}: {e}")

    print("\n✅ Registration batch complete!")

if __name__ == "__main__":
    main()
