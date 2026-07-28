#!/usr/bin/env python3
import json
import csv
import random
import time
import requests
import threading
import concurrent.futures
import fcntl
import openpyxl
from pathlib import Path

# Paths
NEW_CSV_PATH = Path("NEW.csv")
DB_PATH = Path("Names_db.json")
USED_EMAILS_PATH = Path("used_emails.json")
EXCEL_PATH = Path("Names.xlsx")
ACTIVATIONS_JSON_PATH = Path("mails/activations.json")
ACTIVATIONS_JS_PATH = Path("mails/activations.js")

BASE_URL = "https://aileaders.uz"
GMAIL_USER = "qwertyuioplkjhgfdsazxcvbnmhrh@gmail.com"

# Settings
CONCURRENCY = 20
START_INDEX = 4085
TARGET_COUNT = 1400  # Target number of unactivated accounts to clean up & re-register

HEADERS = {
    "accept": "*/*",
    "accept-language": "en-US,en;q=0.9,uz;q=0.8",
    "content-type": "application/json",
    "origin": "https://aileaders.uz",
    "priority": "u=1, i",
    "referer": "https://aileaders.uz/auth/register",
    "sec-ch-ua": "\"Not;A=Brand\";v=\"8\", \"Chromium\";v=\"150\", \"Google Chrome\";v=\"150\"",
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": "\"Linux\"",
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-origin",
    "user-agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36",
}

lock = threading.Lock()

# Load and seed ALL used emails from every source to guarantee NO email is ever reused
used_emails = set()

# 1. From used_emails.json
if USED_EMAILS_PATH.exists():
    try:
        raw_used = json.loads(USED_EMAILS_PATH.read_text())
        for e in raw_used:
            if isinstance(e, str) and e.strip():
                used_emails.add(e.strip().lower())
    except Exception as e:
        print(f"Warning loading used_emails.json: {e}")

# 2. From NEW.csv
if NEW_CSV_PATH.exists():
    try:
        with open(NEW_CSV_PATH, encoding="utf-8") as f:
            r = csv.DictReader(f)
            for row in r:
                em = (row.get("email") or "").strip().lower()
                if em:
                    used_emails.add(em)
    except Exception as e:
        print(f"Warning loading NEW.csv emails: {e}")

print(f"🔒 Initialized used_emails cache with {len(used_emails)} unique historically used emails.")

base_username = GMAIL_USER.split("@")[0]

def generate_dot_alias():
    L = len(base_username)
    while True:
        parts = [base_username[0]]
        for i in range(1, L):
            if random.choice([True, False]):
                parts.append(".")
            parts.append(base_username[i])
        email_addr = "".join(parts) + "@gmail.com"
        email_lower = email_addr.lower()
        with lock:
            if email_lower not in used_emails:
                used_emails.add(email_lower)
                return email_addr

def flush_used_emails():
    with lock:
        try:
            USED_EMAILS_PATH.write_text(json.dumps(list(used_emails), indent=2))
        except Exception as e:
            print(f"Error flushing used emails: {e}")

def first_api_call(session, document, dob):
    url = f"{BASE_URL}/api/public/info/individual"
    params = {"document": document, "dob": dob, "occupation": "student"}
    response = session.post(url, params=params, headers=HEADERS, data="")
    response.raise_for_status()
    return response.json() if response.content else {}

def second_api_call(session, email_addr, document, dob, phone):
    valid_phone = phone if (phone and phone.startswith("+998") and len(phone) == 13) else "+998995337221"
    url = f"{BASE_URL}/api/registration/form"
    payload = {
        "email": email_addr,
        "employment_type": "student",
        "metrika": None,
        "passport": {"document": document, "dob": dob},
        "password": document,
        "phone": valid_phone,
    }
    response = session.post(url, headers=HEADERS, json=payload)
    response.raise_for_status()
    return response.json()

def delete_account(session, document, dob):
    url = f"{BASE_URL}/api/profile/delete-account"
    headers = HEADERS.copy()
    headers["content-type"] = "application/x-www-form-urlencoded"
    headers["referer"] = f"{BASE_URL}/auth/delete_account"
    payload = f"document={document}&dob={dob}"
    try:
        response = session.delete(url, headers=headers, data=payload)
        if response.status_code == 200:
            return response.json() if response.content else {}
    except Exception as e:
        print(f"⚠️ Warning deleting account {document}: {e}")
    return {}

# Memory buffer for results
results_buffer = {}
results_lock = threading.Lock()

def record_registration(student_id, email_addr, password):
    with results_lock:
        results_buffer[student_id] = (email_addr, password)

def flush_all_data():
    with results_lock:
        if not results_buffer:
            return
        current_results = dict(results_buffer)

    # 1. Update NEW.csv
    try:
        lock_path = NEW_CSV_PATH.with_suffix(".csv.lock")
        with open(lock_path, "w") as lock_f:
            fcntl.flock(lock_f, fcntl.LOCK_EX)
            with open(NEW_CSV_PATH, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                rows = list(reader)

            updated_count = 0
            for row in rows[1:]:
                if len(row) > 0 and row[0] in current_results:
                    email_addr, password = current_results[row[0]]
                    while len(row) < 13:
                        row.append("")
                    row[10] = email_addr  # Column K: email
                    row[11] = password    # Column L: password
                    updated_count += 1

            with open(NEW_CSV_PATH, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerows(rows)
            fcntl.flock(lock_f, fcntl.LOCK_UN)
            print(f"💾 Flushed {updated_count} updates to NEW.csv")
    except Exception as e:
        print(f"⚠️ Error flushing NEW.csv: {e}")

    # 2. Update Names_db.json
    try:
        if DB_PATH.exists():
            lock_path = DB_PATH.with_suffix(".json.lock")
            with open(lock_path, "w") as lock_f:
                fcntl.flock(lock_f, fcntl.LOCK_EX)
                db = json.loads(DB_PATH.read_text())
                for item in db:
                    doc = item.get("document")
                    if doc in current_results:
                        item["email"] = current_results[doc][0]
                        item["activated"] = False
                DB_PATH.write_text(json.dumps(db, ensure_ascii=False, indent=2))
                fcntl.flock(lock_f, fcntl.LOCK_UN)
                print(f"💾 Flushed updates to Names_db.json")
    except Exception as e:
        print(f"⚠️ Error flushing Names_db.json: {e}")

    # 3. Update used_emails.json
    flush_used_emails()

    # 4. Update Names.xlsx
    try:
        if EXCEL_PATH.exists():
            excel_lock_path = EXCEL_PATH.with_suffix(".xlsx.lock")
            with open(excel_lock_path, "w") as lock_f:
                fcntl.flock(lock_f, fcntl.LOCK_EX)
                wb = openpyxl.load_workbook(EXCEL_PATH)
                ws = wb["Talabalar"]
                passport_to_row = {}
                for r in range(2, ws.max_row + 1):
                    pass_val = str(ws.cell(row=r, column=2).value or "").strip()
                    if pass_val:
                        passport_to_row[pass_val] = r
                for doc, (email_addr, _) in current_results.items():
                    row_idx = passport_to_row.get(doc)
                    if row_idx:
                        ws.cell(row=row_idx, column=7, value=email_addr)
                wb.save(EXCEL_PATH)
                print("💾 Flushed updates to Names.xlsx")
                fcntl.flock(lock_f, fcntl.LOCK_UN)
    except Exception as e:
        print(f"⚠️ Error flushing Names.xlsx: {e}")

success_counter = 0
counter_lock = threading.Lock()

def worker_task(worker_id, row_index, student_id, full_name, dob, phone):
    global success_counter
    session = requests.Session()

    # Step 1: Delete existing unactivated account on AI Leaders
    print(f"🗑️ [Worker {worker_id}] Deleting existing unactivated account for {student_id} ({full_name})")
    delete_account(session, student_id, dob)
    time.sleep(0.1)

    attempts = 0
    max_attempts = 5
    success = False

    while attempts < max_attempts and not success:
        attempts += 1
        email_addr = generate_dot_alias()

        try:
            # 1. Info check
            first_api_call(session, student_id, dob)
            time.sleep(0.05)

            # 2. Registration form
            resp = second_api_call(session, email_addr, student_id, dob, phone)
            res_code = resp.get("result", {}).get("code")

            if res_code != "ok":
                if res_code == "passport_is_already_in_use":
                    # Delete account again and retry
                    delete_account(session, student_id, dob)
                    time.sleep(0.1)
                    continue
                elif "email" in str(res_code):
                    continue
                else:
                    print(f"❌ [Worker {worker_id}] Student {student_id} ({full_name}) rejected: {res_code}")
                    break

            # 3. Authorization login to get JWT token
            login_resp = session.post(f"{BASE_URL}/api/authorization/login", headers={
                **HEADERS,
                "referer": f"{BASE_URL}/auth/login",
            }, json={
                "login": email_addr,
                "password": student_id
            })
            login_resp.raise_for_status()
            token = login_resp.json().get("content", {}).get("token")

            if not token:
                continue

            # 4. Trigger email verify endpoint (dispatches activation email)
            verify_resp = session.post(f"{BASE_URL}/api/profile/verify-email?email={email_addr}", headers={
                **HEADERS,
                "Authorization": f"Bearer {token}"
            })
            verify_resp.raise_for_status()

            record_registration(student_id, email_addr, student_id)

            with counter_lock:
                success_counter += 1
                current_cnt = success_counter

            print(f"✅ [{current_cnt}/{TARGET_COUNT}] Re-registered {student_id} ({full_name}) -> {email_addr}")
            success = True

            if current_cnt % 50 == 0:
                flush_all_data()

        except Exception as e:
            time.sleep(0.2)

def load_activations():
    activated_emails = set()
    activated_docs = set()

    # 1. Load activations.json
    if ACTIVATIONS_JSON_PATH.exists():
        try:
            data = json.loads(ACTIVATIONS_JSON_PATH.read_text())
            for item in data:
                if item.get("email"):
                    activated_emails.add(item["email"].strip().lower())
                if item.get("document"):
                    activated_docs.add(item["document"].strip())
        except Exception as e:
            print(f"Warning reading activations.json: {e}")

    # 2. Load activations.js
    if ACTIVATIONS_JS_PATH.exists():
        try:
            content = ACTIVATIONS_JS_PATH.read_text().replace("module.exports =", "").strip()
            if content.endswith(";"):
                content = content[:-1].strip()
            data = json.loads(content)
            for item in data:
                if item.get("email"):
                    activated_emails.add(item["email"].strip().lower())
                if item.get("document"):
                    activated_docs.add(item["document"].strip())
        except Exception as e:
            print(f"Warning reading activations.js: {e}")

    return activated_emails, activated_docs

def main():
    activated_emails, activated_docs = load_activations()
    print(f"📩 Loaded {len(activated_emails)} activated emails and {len(activated_docs)} activated documents from activations.js / json.")

    # Seed activated emails into used_emails set
    for em in activated_emails:
        used_emails.add(em)
    flush_used_emails()

    # Load NEW.csv
    with open(NEW_CSV_PATH, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    # Load Names_db.json for DOB & phone lookup
    db_map = {}
    if DB_PATH.exists():
        try:
            db_entries = json.loads(DB_PATH.read_text())
            for item in db_entries:
                doc = item.get("document")
                if doc:
                    db_map[doc] = item
        except Exception as e:
            print(f"Error loading Names_db.json: {e}")

    # Build candidates list starting from START_INDEX (4085)
    candidates = []
    for idx in range(START_INDEX, len(rows)):
        r = rows[idx]
        student_id = (r.get("student_id") or "").strip()
        full_name = (r.get("full_name") or "").strip()
        email = (r.get("email") or "").strip().lower()
        cert = (r.get("certificate_url") or "").strip()

        # Strict rules requested by user:
        # 1. NEVER delete/touch an account that has a certificate in NEW.csv
        if cert:
            continue

        # 2. Check if email or document is in activations.js / json (if activated, skip)
        if email and (email in activated_emails or student_id in activated_docs):
            continue

        # 3. Candidate for account termination & re-registration with a new unique email
        if student_id:
            db_item = db_map.get(student_id, {})
            dob = db_item.get("dob")
            phone = db_item.get("phone")
            if dob:
                candidates.append((idx, student_id, full_name, dob, phone))

        if len(candidates) >= TARGET_COUNT:
            break

    print(f"📋 Found {len(candidates)} candidates for account termination and re-registration (starting at index {START_INDEX}).")

    start_time = time.time()

    with concurrent.futures.ThreadPoolExecutor(max_workers=CONCURRENCY) as executor:
        futures = []
        for worker_id, (idx, student_id, full_name, dob, phone) in enumerate(candidates, 1):
            f = executor.submit(worker_task, worker_id, idx, student_id, full_name, dob, phone)
            futures.append(f)

        concurrent.futures.wait(futures)

    flush_all_data()
    elapsed = time.time() - start_time
    print(f"🎉 Cleanup & Re-registration finished in {elapsed:.2f} seconds!")
    print(f"✅ Total re-registered in this run: {success_counter}")

if __name__ == "__main__":
    main()
