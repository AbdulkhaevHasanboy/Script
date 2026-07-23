import json
import openpyxl
from pathlib import Path

# Config
EXCEL_PATH = Path("Names.xlsx")
RESULTS_PATH = Path("results.json")

def main():
    if not RESULTS_PATH.exists():
        print("Error: results.json not found.")
        return
    if not EXCEL_PATH.exists():
        print("Error: Names.xlsx not found.")
        return

    # Load results.json
    results = json.loads(RESULTS_PATH.read_text())
    print(f"Loaded {len(results)} registered students from results.json.")

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Talabalar"]
    except Exception as e:
        print(f"Error opening Excel file: {e}")
        return

    # Ensure headers in F and G
    ws.cell(row=1, column=6, value="Invite Link")
    ws.cell(row=1, column=7, value="Email")

    # Map passport to sheet row
    passport_to_row = {}
    for r in range(3, ws.max_row + 1):
        pass_val = str(ws.cell(row=r, column=2).value or "").strip()
        if pass_val:
            passport_to_row[pass_val] = r

    # Write emails from results.json
    written_emails = 0
    for entry in results:
        passport = entry.get("document")
        email_addr = entry.get("email")
        if passport and email_addr:
            row_idx = passport_to_row.get(passport)
            if row_idx:
                ws.cell(row=row_idx, column=7, value=email_addr)
                written_emails += 1

    wb.save(EXCEL_PATH)
    print(f"Success! Restored {written_emails} email addresses to Column G of Names.xlsx.")

if __name__ == "__main__":
    main()
