#!/usr/bin/env python3
import json
import csv
import fcntl
import openpyxl
from pathlib import Path

NEW_CSV_PATH = Path("NEW.csv")
EXCEL_PATH = Path("Names.xlsx")
COURSERA_JSON_PATH = Path("mails/coursera/coursera_mails.json")

def main():
    if not COURSERA_JSON_PATH.exists():
        print(f"❌ {COURSERA_JSON_PATH} not found.")
        return

    records = json.loads(COURSERA_JSON_PATH.read_text())
    print(f"📋 Loaded {len(records)} Coursera records.")

    doc_map = {r["document"].strip(): r["invitationUrl"] for r in records if r.get("document") and r.get("invitationUrl")}
    email_map = {r["email"].strip().lower(): r["invitationUrl"] for r in records if r.get("email") and r.get("invitationUrl")}

    # 1. Update NEW.csv
    updated_csv_count = 0
    if NEW_CSV_PATH.exists():
        lock_path = NEW_CSV_PATH.with_suffix(".csv.lock")
        with open(lock_path, "w") as lock_f:
            fcntl.flock(lock_f, fcntl.LOCK_EX)
            with open(NEW_CSV_PATH, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                rows = list(reader)

            header = rows[0]
            # Column C: invit_url (index 2), Column M: certificate_url (index 12)
            for row in rows[1:]:
                if not row:
                    continue
                doc = row[0].strip() if len(row) > 0 else ""
                email_addr = row[10].strip().lower() if len(row) > 10 else ""
                cert_url = row[12].strip() if len(row) > 12 else ""

                # Rule: ONLY update if Column M (certificate_url) is empty
                if not cert_url:
                    inv_url = doc_map.get(doc) or email_map.get(email_addr)
                    if inv_url:
                        while len(row) < 13:
                            row.append("")
                        row[2] = inv_url  # Column C: invit_url
                        updated_csv_count += 1

            with open(NEW_CSV_PATH, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerows(rows)
            fcntl.flock(lock_f, fcntl.LOCK_UN)

        print(f"💾 Updated {updated_csv_count} rows in NEW.csv (Column C invit_url).")

    # 2. Update Names.xlsx
    updated_excel_count = 0
    if EXCEL_PATH.exists():
        excel_lock_path = EXCEL_PATH.with_suffix(".xlsx.lock")
        with open(excel_lock_path, "w") as lock_f:
            fcntl.flock(lock_f, fcntl.LOCK_EX)
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb["Talabalar"]

            # Column 2: Passport, Column 6: Invite Link, Column 7: Email, Column 8: Certificate URL
            for r in range(2, ws.max_row + 1):
                pass_val = str(ws.cell(row=r, column=2).value or "").strip()
                email_val = str(ws.cell(row=r, column=7).value or "").strip().lower()
                cert_val = str(ws.cell(row=r, column=8).value or "").strip()

                # Rule: ONLY update if Certificate URL (Column 8) is empty
                if not cert_val:
                    inv_url = doc_map.get(pass_val) or email_map.get(email_val)
                    if inv_url:
                        ws.cell(row=r, column=6, value=inv_url)
                        updated_excel_count += 1

            wb.save(EXCEL_PATH)
            fcntl.flock(lock_f, fcntl.LOCK_UN)

        print(f"💾 Updated {updated_excel_count} rows in Names.xlsx (Column F Invite Link).")

if __name__ == "__main__":
    main()
